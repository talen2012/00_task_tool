import pandas as pd
from openpyxl import load_workbook
import sys

# -*- coding: utf-8 -*-
# @Time    : 2025/8/29 13:45
# @Author  : talen
# @File    : category_id_gen.py

"""
生成各列的序号，并根据此序列号生成对应的编号
编号时考虑当前分类类及其所有上级分类，当不完全相同时，编号不同
"""

def generate_class_orders_dicts(file_path):
    """
    读取指定Excel文件的“生态基础表明细”sheet，生成六个字典
    每个字典的键是该级分类及其所有上级分类组成的元组，值是该分类的编号
    返回包含六个字典的列表
    """
    # Open the Excel file and read the "生态基础表明细" sheet
    try:
        df = pd.read_excel(file_path, sheet_name="生态基础表明细", header=1) # header=1表示表头在第二行（索引从0开始）
    except Exception as e:
        print(f"Error reading file: {e}")
        return

    # Ensure the sheet has 6 columns
    if df.shape[1] != 6:
        print("The sheet must have exactly 6 columns.")
        return

    # Initialize a dictionary to store unique classifications and their numbers
    class_orders_dicts = [{} for _ in range(6)] # 生成6个空字典
    current_orders = [1] * 6 #   序号从1开始，每级独立

    # Iterate through each row in the sheet
    import sys
    for idx, row in df.iterrows():
        for level in range(6):
            if pd.isna(row[level]) or row[level] == "/":
                print(f"Error: Invalid value at row {idx + 2} (Excel 行号), 内容: {row.tolist()}")
                sys.exit(1)
            # 各级分类及其所有上级分类组成的元组不重复时，编号递增
            classification = tuple(row[:level + 1])

            # Assign a number if the classification is new
            if classification not in class_orders_dicts[level]:
                class_orders_dicts[level][classification] = current_orders[level]
                current_orders[level] += 1
    return class_orders_dicts    

def generate_class_orders(class_orders_dicts, file_path):
    """
    生成各级分类的序号，并写入Excel文件的“各级序号”sheet，从第5行开始写入
    每一级分类占据两列，第一列是分类名称，第二列是对应的序号
    各级分类及序号独立、连续，不受其他级分类影响
    """
    if not class_orders_dicts:
        print("Failed to generate class orders.")
        return

    # 每一级单独生成一个DataFrame，每级只填自己那两列，其余列全为None，最后按行拼接
    level_dfs = []
    for level in range(6):
        rows = []
        for classification, number in class_orders_dicts[level].items():
            row = [None] * 2
            row[0] = classification[level]
            row[1] = number
            rows.append(row)
        level_df = pd.DataFrame(rows)
        level_dfs.append(level_df)
    output_df = pd.concat(level_dfs, axis=1, ignore_index=True) # axis=1表示按列拼接，ignore_index=True表示重新索引
    # df.where(条件, 替换值) 条件为True的位置保留原值，条件为False的位置用“替换值”填充
    output_df = output_df.where(pd.notnull(output_df), None) # 将NaN替换为None

    # 用openpyxl将output_df从第5行开始写入“各类序号”sheet，不影响前四行内容和公式
    # pandas的ExcelWriter不支持从指定行写入，只能覆盖整个sheet

    wb = load_workbook(file_path)
    if "各级序号" in wb.sheetnames:
        ws = wb["各级序号"]
    else:
        ws = wb.create_sheet("各级序号")

    # 清除第5行及之后的内容（可选，防止旧数据残留）
    max_row = ws.max_row
    if max_row >= 5:
        for row in ws.iter_rows(min_row=5, max_row=max_row):
            for cell in row:
                cell.value = None

    # 将output_df写入第5行及之后
    for i, row in enumerate(output_df.values, start=5): # output_df.values返回的是二维数组，数据类型是numpy.ndarray
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    wb.save(file_path)


def generate_class_id(class_orders_dicts, file_path, sheet_name):
    """
    在已有各级分类序号的基础上，对于一份完整的行业列+五级分类表，生成每个分类项的id
    行业列id由1位数字构成
    这些id由分类级别+规定位数的序号组成，如一级分类id=101，二级分类id=2001，三级分类id=3001
    各级分类序号分别规定两位，三位，三位，三位，四位数字
    """
    if not class_orders_dicts:
        print("Failed to generate class ids.")
        return
    # 用openpyxl读取指定sheet,依据分类名称信息，生成对应的分类id并写入相应单元格
    wb = load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"Sheet {sheet_name} does not exist.")
        return
    
    if ws.max_row < 2:
        print(f"Sheet {sheet_name} does not have enough rows.")
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row): # min_row=2表示从第二行开始
        # row是一个元组，包含该行的所有单元格对象，对象类型是openpyxl.cell.cell.Cell
        for level in range(6):
            if row[level * 2].value is None or row[level * 2].value == "/":
                row[level * 2].value = row[(level - 1) * 2].value if level > 0 else None
            # 获取当前分类及其所有上级分类组成的元组（每级分类占两列，取奇数列的值）
            classification = tuple(cell.value for cell in row[:level * 2 + 1:2])
            # 查找对应的序号
            if classification in class_orders_dicts[level]:
                order = class_orders_dicts[level][classification]
                # 生成对应的id（可根据实际需求调整位数规则）
                if level == 0:
                    id = order  # 一级分类id=1,2...
                elif level == 1:
                    id = 100 + order  # 二级分类id=201,202...
                elif level == 2:
                    id = 2000 + order  # 三级分类id=2001,2002...
                elif level == 3:
                    id = 3000 + order  # 四级分类id=3001,3002...
                elif level == 4:
                    id = 4000 + order  # 五级分类id=4001,4002...
                elif level == 5:
                    id = 50000 + order  # 六级分类id=601,602...

                row[level * 2 + 1].value = id
            else:
                row[level * 2 + 1].value = "/"  # 如果找不到对应的序号，写入"/"

    wb.save(file_path)


def match_self_with_father_id(file_path):
    '''
    将六级分类编号数据转换为一列，并写出对应的父级编号
    '''
    # 读取指定sheet
    # header=0表示表头在第一行（索引从0开始）
    df = pd.read_excel(file_path, sheet_name="全量_编码", header=0) 

    # 结果DataFrame
    result = pd.DataFrame(columns=["field", "field_id", "parent_id", "level"])

    # 处理六级分类
    # 先全部堆叠，不去重
    num_levels = 6
    for level in range(num_levels):
        name_col = df.iloc[:, level * 2]
        id_col = df.iloc[:, level * 2 + 1] # iloc按位置索引列
        if level == 0:
            parent_id = [None] * len(df) # 一级分类没有父类
        else:
            parent_id = df.iloc[:, (level - 1) * 2 + 1]
        temp_df = pd.DataFrame({
            "field": name_col,
            "field_id": id_col,
            "parent_id": parent_id,
            "level": [level] * len(df)
        })
        result = pd.concat([result, temp_df], ignore_index=True)



    # 去除空行
    result = result.dropna(subset=["field_id"])
    # 删除feild_id为“/”的数据行
    result = result[result["field_id"] != "/"]
    # 去除类别和编号同时重复的行（保留首次出现的编号）
    result = result.drop_duplicates(subset=["field","field_id"], keep="first") # keep="first"表示保留第一次出现的行

    # 保存结果
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        result.to_excel(writer, sheet_name="五级分类_编码", index=False)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python class_number_gen.py <file_path>")
    else:
        class_orders_dicts = generate_class_orders_dicts(sys.argv[1])
        generate_class_orders(class_orders_dicts, sys.argv[1])
        generate_class_id(class_orders_dicts, sys.argv[1], "全量_编码")
        generate_class_id(class_orders_dicts, sys.argv[1], "有厂家部分_编码")
        match_self_with_father_id(sys.argv[1])
