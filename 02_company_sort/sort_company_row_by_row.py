import openpyxl

# @Time    : 2025/09/05 17:26
# @Author  : talen
# @File    : sort_company_row_by_row.py

def sort_company_in_row(companies_info):
    """
    对一行内的公司信息进行排序，单个公司信息由名称、生态类型和分数构成
    一行有多个公司
    :param companies_info: 待排序的公司数据（一维列表）
    :return: 排序后的公司数据
    """
    # 列表元素不是三的倍数，补齐最后一组
    if len(companies_info) % 3 != 0:
        companies_info.extend([""] * (3 - len(companies_info) % 3))
    # 将公司信息按每三个元素一组进行分割
    companies_info_split = [companies_info[i:i + 3] for i in range(0, len(companies_info), 3)]
    # 先按照生态类型，即第二列排列，排列顺序为 自有公司、西安生态、数博会
    # 每种生态类型内部，再按照得分降序排列，考虑得分为空或“/”，排在最后
    companies_info_split.sort(
        key = lambda x: (
            {"自有公司": 2, "西安生态": 1, "数博会": 0}.get(x[1], -1),  # 按生态类型
            float(x[2]) if x[2] not in (None, "", "/", "/ ") else float('-inf') # 按得分降序
        ),
        reverse=True
    )
    # 将排序后的公司信息重新拼接为一维列表
    companies_info = [item for sublist in companies_info_split for item in sublist]

    return companies_info

target_file = "02_company_sort\\test.xlsx"
target_sheet = "生态基础表明细-能力"
wb = openpyxl.load_workbook(target_file)
ws = wb[target_sheet]
print(f"读取文件 {target_file} 中的工作表 {target_sheet} 成功")
# openpyxl对数组公式不支持，会处理为字符串，带{}包围，导致处理出错

# 找到第一个名为“生态合作伙伴清单”的列,从1开始计数
first_company_col_no = None
col_labels = [cell.value for cell in ws[1]]  # 假设第一行是表头，ws从1开始计数
for i, label in enumerate(col_labels):
    if label == "生态合作伙伴清单":
        first_company_col_no = i
        break
print("公司信息排序中...")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row): # 从第二行开始遍历
    companies_info = [cell.value for cell in row[first_company_col_no:]] # row是一个元组，从0开始计数
    sorted_companies_info = sort_company_in_row(companies_info)

    # 将排序后的公司信息写回单元格
    for col, company_info in enumerate(sorted_companies_info, start=first_company_col_no):
        row[col].value = company_info
print("排序完成！")
# 保存修改后的Excel文件
wb.save(target_file)