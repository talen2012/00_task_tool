import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
import numpy as np
import os
import threading
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
import re
import difflib
import traceback

# @Time    : 2025/12/04/10:43
# @Author  : talen
# @File    : bid_analysis_tool.py
class BidAnalysisApp:
    def __init__(self, root):
        self.root = root
        self.root.title("中标信息分析工具")
        self.root.geometry("800x600")

        # 变量
        self.source_file_path = tk.StringVar() # tk.StringVar()是用来创建一个字符串变量，用于绑定到Tkinter的输入控件上，以便动态更新和获取用户输入的值。
        self.config_file_path = tk.StringVar()
        self.project_file_path = tk.StringVar()
        self.company_file_path = tk.StringVar()
        self.status_msg = tk.StringVar(value="就绪")

        # 默认配置路径（如果存在）
        default_config = os.path.join(os.path.dirname(__file__), "关键词配置表.xlsx")
        if os.path.exists(default_config):
            self.config_file_path.set(default_config)

        # 定义目标列（统一Schema）
        self.target_columns = [
            "序号", "省份", "市", "区县", "中标年份", "中标月份", "中标时间", "招标类型", 
            "项目名称", "招标单位", "中标单位", "中标金额（万元）", 
            "中标厂商类型", "项目所属行业", "备注", "所属行业（ICT）", 
            "所属业务类型（ICT）", "公告内容（ICT）", "行业（省公司）", "数据来源", "修改时间"
        ]
        # 用于保证同一时间只有一个分析线程在运行
        self.analysis_lock = threading.Lock() # 互斥锁

        self._create_ui()

    def _create_ui(self):
        # 1. 文件选择区域
        frame_files = tk.LabelFrame(self.root, text="文件选择", padx=10, pady=10)
        frame_files.pack(fill="x", padx=10, pady=5) # pack方法用于将控件放置在窗口中，并设置填充和边距。fill="x"表示水平填充，padx和pady分别表示水平和垂直方向的内边距。默认情况下，pack方法会将控件放置在父容器的顶部，并根据需要调整大小。

        # 数据源文件
        tk.Label(frame_files, text="项目源数据(Excel):").grid(row=0, column=0, sticky="w") # sticky="w"表示标签左对齐，grid方法用于将控件放置在网格布局中的指定行和列。
        tk.Entry(frame_files, textvariable=self.source_file_path, width=80).grid(row=0, column=1, padx=5)
        tk.Button(frame_files, text="浏览...", command=self.select_source_file).grid(row=0, column=2)

        # 配置文件
        tk.Label(frame_files, text="关键词配置(Excel):").grid(row=1, column=0, sticky="w")
        tk.Entry(frame_files, textvariable=self.config_file_path, width=80).grid(row=1, column=1, padx=5) 
        tk.Button(frame_files, text="浏览...", command=self.select_config_file).grid(row=1, column=2)

        # 项目汇总文件
        tk.Label(frame_files, text="项目汇总文件(Excel):").grid(row=2, column=0, sticky="w")
        tk.Entry(frame_files, textvariable=self.project_file_path, width=80).grid(row=2, column=1, padx=5)
        tk.Button(frame_files, text="浏览...", command=self.select_project_file).grid(row=2, column=2)

        # 公司汇总文件
        tk.Label(frame_files, text="公司汇总文件(Excel):").grid(row=3, column=0, sticky="w")
        tk.Entry(frame_files, textvariable=self.company_file_path, width=80).grid(row=3, column=1, padx=5)
        tk.Button(frame_files, text="浏览...", command=self.select_company_file).grid(row=3, column=2)

        # 2. 操作区域
        frame_actions = tk.Frame(self.root, padx=10, pady=10)
        frame_actions.pack(fill="x", padx=10)

        tk.Button(frame_actions, text="项目汇总", command=self.start_project_analysis_thread, bg="#4CAF50", fg="white", font=("Microsoft YaHei", 10, "bold")).pack(side="left", padx=5)
        tk.Button(frame_actions, text="公司汇总", command=self.start_company_analysis_thread, bg="#4CAF50", fg="white", font=("Microsoft YaHei", 10, "bold")).pack(side="left", padx=50) # padx表示按钮之间的水平间距
        tk.Label(frame_actions, textvariable=self.status_msg, fg="blue").pack(side="left", padx=0)

        # 3. 日志区域
        frame_log = tk.LabelFrame(self.root, text="运行日志", padx=10, pady=10)
        frame_log.pack(fill="both", expand=True, padx=10, pady=5) # expand=True表示日志区域会随着窗口大小调整而扩展，占据更多空间。 fill="both"表示水平和垂直方向都填充。pady表示距离上一个控件的垂直距离。

        self.log_text = scrolledtext.ScrolledText(frame_log, height=20)
        self.log_text.pack(fill="both", expand=True)

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END) # see方法用于将文本框的视图滚动到最后一行，以确保最新的日志消息始终可见。

    def select_source_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            self.source_file_path.set(filename)

    def select_config_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            self.config_file_path.set(filename)

    def select_project_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            self.project_file_path.set(filename)

    def select_company_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filename:
            self.company_file_path.set(filename)

    def _start_thread_with_lock(self, target, start_msg="正在运行分析..."):
        """尝试获取分析锁并在单独线程中运行 target；若已被占用则提示用户。"""
        if not self.analysis_lock.acquire(blocking=False): 
            messagebox.showwarning("提示", "已有分析在运行，请稍后再试。")
            return

        def _worker():
            try:
                self.status_msg.set(start_msg)
                target()
            finally:
                try:
                    self.analysis_lock.release()
                except RuntimeError:
                    pass # 锁已被释放，无需操作
                
        thread = threading.Thread(target=_worker)
        thread.daemon = True
        thread.start()

    def start_project_analysis_thread(self):
        if not self.source_file_path.get():
            messagebox.showwarning("提示", "请选择源数据文件")
            return
        if not self.config_file_path.get():
            messagebox.showwarning("提示", "请选择关键词配置文件")
            return
        if not self.project_file_path.get():
            messagebox.showwarning("提示", "请选择项目汇总文件")
            return
        
        # 通过锁保证与公司汇总不并发运行
        self._start_thread_with_lock(self.run_project_analysis_workflow, start_msg="正在整理项目信息...")

    def start_company_analysis_thread(self):
        if not self.company_file_path.get():
            messagebox.showwarning("提示", "请选择公司汇总文件")
            return
        # 通过锁保证与项目汇总不并发运行
        self._start_thread_with_lock(self.run_company_analysis_workflow, start_msg="正在汇总公司信息...")

    def _split_region_direct(self, region):
        """拆分区域字符串为省、市、区县"""
        if pd.isna(region) or region.strip() == "":
            return "未知", "未知", "未知"
        
        region = region.strip()
        p_idx = region.find("省")  # 省的位置（无则返回-1）
        c_idx = region.find("市")  # 市的位置（无则返回-1）
        d_idx = region.find("区")  # 区的位置（无则返回-1）
        
        # 拆分省份：从开头到“省”字（不含“省”），无“省”则未知
        province = region[:p_idx] if p_idx != -1 else "未知"
        
        # 拆分市：从“省”后到“市”字（不含“市”），无“省”则从开头找“市”，无“市”则未知
        if p_idx != -1 and c_idx != -1 and c_idx > p_idx:
            city = region[p_idx+1:c_idx] 
        elif c_idx != -1:  # 极端情况：无“省”但有“市”（如“西安市雁塔区”）
            city = region[:c_idx]
        else:
            city = "未知"
        
        # 拆分区县：从“市”后到“区”字（包含“区”），无“市”则从“省”后找“区”，无“区”则未知
        if c_idx != -1 and d_idx != -1 and d_idx > c_idx:
            district = region[c_idx+1:d_idx+1]
        elif d_idx != -1:  # 极端情况：仅“区”（如“雁塔区”）
            district = region[:d_idx+1]
        else:
            district = "未知"
        
        return province, city, district

    def _get_keyword_best_match(self, text, keyword_map, default_val):
        if not keyword_map:
            return default_val
        hit_counts = {k: 0 for k in keyword_map.keys()}
        for k, kws in keyword_map.items():
            for kw in kws:
                if kw in text:
                    hit_counts[k] += 1
        max_hits = max(hit_counts.values())
        if max_hits > 0:
            for k, count in hit_counts.items():
                if count == max_hits:
                    return k
        return default_val
    
    def _analyze_vendor(self, row):
            if '中标单位' not in row or pd.isna(row.get('中标单位', "")):
                return "未知"
            return self._get_keyword_best_match(str(row['中标单位']), self.vendor_map, "其他厂商")

    def _analyze_industry(self, row):
        cols = ['项目名称', '招标单位', '所属行业（ICT）', '所属业务类型（ICT）', '行业（省公司）']
        text = ""
        for c in cols:
            text += str(row.get(c, "")) + " "
        return self._get_keyword_best_match(text, self.industry_map, "未分类")
    
    # 将可能含有多个中标单位的字段拆分为多行。常见分隔符包括英文/中文逗号、顿号和分号。
    def _split_bidders_field(self, val):
        if pd.isna(val):
            return [val]
        parts = re.split(r'[，,、;；]+', str(val))
        parts = [p.strip() for p in parts if p and p.strip()]
        return parts if parts else [val] # 确保至少返回一个原始值

    def _process_ict_data(self, source_path):
        self.log("正在使用 [ICT标局] 格式进行处理项目记录...")
        try:
            df_source = pd.read_excel(source_path, header=1)  # header=1表示第二行是列名
        except Exception as e:
            self.log(f"错误: 读取源数据失败 - {e}")
            self.status_msg.set("源数据读取失败")
            return None

        df_new_bid = pd.DataFrame(columns=self.target_columns)
        
        try:
            # 按照列名映射关系提取数据
            df_new_bid["省份"], df_new_bid["市"], df_new_bid["区县"] = zip(*df_source["区域"].apply(self._split_region_direct)) # apply方法返回值类型为一个Series，*将Series解压为多个元组，zip将每个元组同一位置的元素组合成一个新的元组
            ## 日期格式转换
            df_source["发布时间"] = pd.to_datetime(df_source["发布时间"], errors='coerce')  # errors='coerce'参数用于将无法解析的日期转换为NaT（Not a Time），避免程序报错。
            df_new_bid["中标月份"] = df_source["发布时间"].dt.strftime("%Y%m") # %m表示月份是两位数格式，不足补0
            df_new_bid["中标年份"] = df_source["发布时间"].dt.strftime("%y年") # %y表示两位数年份
            df_new_bid['中标时间'] = df_source['发布时间'].dt.strftime("%Y-%m-%d")
            
            df_new_bid['招标类型'] = df_source['公告类型']
            df_new_bid['项目名称'] = df_source['公告名称']
            df_new_bid['招标单位'] = df_source['招标单位']
            df_new_bid['中标单位'] = df_source['中标单位']
            df_new_bid['中标金额（万元）'] = df_source['中标金额（万元）']
            df_new_bid['所属行业（ICT）'] = df_source['所属行业']
            df_new_bid['所属业务类型（ICT）'] = df_source['所属业务类型']
            df_new_bid['公告内容（ICT）'] = df_source['公告内容']
            # 标记数据来源
            df_new_bid['数据来源'] = "ICT标局"

            # 构造列表列并 explode（向量化，效率高）
            df_new_bid['中标单位_list'] = df_new_bid['中标单位'].apply(self._split_bidders_field)
            # 记录每行原始中标单位个数，以便拆分后标记分包
            df_new_bid['中标单位_count'] = df_new_bid['中标单位_list'].apply(len)
            # explode依据列表列拆分行，reset_index重置索引,drop=True表示删除旧索引列
            df_new_bid = df_new_bid.explode('中标单位_list').reset_index(drop=True) 
            df_new_bid['中标单位'] = df_new_bid['中标单位_list']

            # 标记分包：若某原始行拆分出多个中标单位，则拆分后的所有子行在备注列写入 '分包'
            df_new_bid['备注'] = df_new_bid['备注'].fillna('')
            df_new_bid.loc[df_new_bid['中标单位_count'] > 1, '备注'] = '分包' # loc用于基于条件、标签名选择并赋值

            # 清理临时列
            df_new_bid.drop(columns=['中标单位_list', '中标单位_count'], inplace=True) # inplace=True表示在原DataFrame上进行修改，不返回新的DataFrame，drop用于删除指定的列或行。

            # 计算基于单行中标单位的分析列
            df_new_bid['中标厂商类型'] = df_new_bid.apply(self._analyze_vendor, axis=1)
            df_new_bid['项目所属行业'] = df_new_bid.apply(self._analyze_industry, axis=1)

            self.log(f"共处理 {len(df_new_bid)} 条记录 (含拆分后的中标单位行数)")
        except Exception as e:
            self.log(f"错误: 处理源数据失败，请检查sheet名称和列名 - {e}")
            self.status_msg.set("源数据处理失败")
            return None
        return df_new_bid

    def _process_province_data(self, source_path):
        self.log("正在使用 [省公司] 格式进行处理项目记录...")
        try:
            sheet_names = pd.ExcelFile(source_path).sheet_names
            target_sheet = next((sheet for sheet in sheet_names if "中标" in sheet), None) # Next函数用于从可迭代对象中获取第一个满足条件的元素。如果没有找到满足条件的元素，则返回None。
            if not target_sheet:
                self.log("错误: 未找到名称包含 '中标' 的Sheet")
                self.status_msg.set("源数据读取失败")
                return None
            df_source = pd.read_excel(source_path, sheet_name=target_sheet) 
        except Exception as e:
            self.log(f"错误: 读取源数据失败 - {e}")
            self.status_msg.set("源数据读取失败")
            return None

        df_new_bid = pd.DataFrame(columns=self.target_columns)
        
        try:
            # 按照列名映射关系提取数据
            df_new_bid["省份"] = df_source['省份']
            df_new_bid["市"] = df_source['市']
            df_new_bid["区县"] = df_source['区县']
            df_new_bid['中标月份'] = df_source['中标月份']
            df_new_bid['中标时间'] = df_source['中标时间']
            df_new_bid['中标年份'] = pd.to_datetime(df_new_bid['中标时间'], errors='coerce').dt.strftime("%y年")
            df_new_bid['招标类型'] = df_source['招标类型']
            df_new_bid['项目名称'] = df_source['项目名称']
            df_new_bid['招标单位'] = df_source['招标单位']
            df_new_bid['中标单位'] = df_source['中标公司'] if '中标公司' in df_source.columns else df_source['中标单位']
            df_new_bid['中标金额（万元）'] = df_source['中标金额'] / 10000
            df_new_bid['行业（省公司）'] = df_source['行业']
            # 标记数据来源
            df_new_bid['数据来源'] = "省公司"

            # 构造列表列并 explode（向量化，效率高）
            df_new_bid['中标单位_list'] = df_new_bid['中标单位'].apply(self._split_bidders_field)
            # 记录每行原始中标单位个数，以便拆分后标记分包
            df_new_bid['中标单位_count'] = df_new_bid['中标单位_list'].apply(len)
            # explode依据列表列拆分行，reset_index重置索引,drop=True表示删除旧索引列
            df_new_bid = df_new_bid.explode('中标单位_list').reset_index(drop=True) 
            df_new_bid['中标单位'] = df_new_bid['中标单位_list']

            # 标记分包：若某原始行拆分出多个中标单位，则拆分后的所有子行在备注列写入 '分包'
            
            df_new_bid['备注'] = df_new_bid['备注'].fillna('')
            df_new_bid.loc[df_new_bid['中标单位_count'] > 1, '备注'] = '分包' # loc用于基于条件、标签名选择并赋值

            # 清理临时列
            df_new_bid.drop(columns=['中标单位_list', '中标单位_count'], inplace=True) # inplace=True表示在原DataFrame上进行修改，不返回新的DataFrame，drop用于删除指定的列或行。

            # 计算基于单行中标单位的分析列
            df_new_bid['中标厂商类型'] = df_new_bid.apply(self._analyze_vendor, axis=1)
            df_new_bid['项目所属行业'] = df_new_bid.apply(self._analyze_industry, axis=1)

            self.log(f"共处理 {len(df_new_bid)} 条记录 (含拆分后的中标单位行数)")
        except Exception as e:
            self.log(f"错误: 处理源数据失败，请检查sheet名称和列名 - {e}")
            self.status_msg.set("源数据处理失败")
            return None
        return df_new_bid

    def _process_shushuo_data(self, source_path):
        self.log("正在使用 [数说123] 格式进行处理项目记录...")
        try:
            df_source = pd.read_excel(source_path, sheet_name='中标项目') 
        except Exception as e:
            self.log(f"错误: 读取源数据失败 - {e}")
            self.status_msg.set("源数据读取失败")
            return None

        df_new_bid = pd.DataFrame(columns=self.target_columns)
        
        try:
            # 按照列名映射关系提取数据
            df_new_bid["市"] = df_source['市']
            df_new_bid["区县"] = df_source['区/县']
            df_new_bid['省份'] = "陕西"
            df_new_bid['中标年份'] = df_source['中标年份']
            df_new_bid['中标月份'] = df_source['中标月份']
            df_new_bid['招标类型'] = "中标公告"
            df_new_bid['项目名称'] = df_source['项目名称']
            df_new_bid['招标单位'] = df_source['招采单位']
            df_new_bid['中标单位'] = df_source['中标公司']
            df_new_bid['中标金额（万元）'] = df_source['中标金额']
            df_new_bid['公告内容（ICT）'] = df_source['项目建设内容']
           # 标记数据来源
            df_new_bid['数据来源'] = "数说123"

            # 构造列表列并 explode（向量化，效率高）
            df_new_bid['中标单位_list'] = df_new_bid['中标单位'].apply(self._split_bidders_field)
            # 记录每行原始中标单位个数，以便拆分后标记分包
            df_new_bid['中标单位_count'] = df_new_bid['中标单位_list'].apply(len)
            # explode依据列表列拆分行，reset_index重置索引,drop=True表示删除旧索引列
            df_new_bid = df_new_bid.explode('中标单位_list').reset_index(drop=True) 
            df_new_bid['中标单位'] = df_new_bid['中标单位_list']

            # 标记分包：若某原始行拆分出多个中标单位，则拆分后的所有子行在备注列写入 '分包'
            df_new_bid['备注'] = df_new_bid['备注'].fillna('')
            df_new_bid.loc[df_new_bid['中标单位_count'] > 1, '备注'] = '分包' # loc用于基于条件、标签名选择并赋值

            # 清理临时列
            df_new_bid.drop(columns=['中标单位_list', '中标单位_count'], inplace=True) # inplace=True表示在原DataFrame上进行修改，不返回新的DataFrame，drop用于删除指定的列或行。

            # 计算基于单行中标单位的分析列
            df_new_bid['中标厂商类型'] = df_new_bid.apply(self._analyze_vendor, axis=1)
            df_new_bid['项目所属行业'] = df_new_bid.apply(self._analyze_industry, axis=1)

            self.log(f"共处理 {len(df_new_bid)} 条记录 (含拆分后的中标单位行数)")
        except Exception as e:
            self.log(f"错误: 处理源数据失败，请检查sheet名称和列名 - {e}")
            self.status_msg.set("源数据处理失败")
            return None
        return df_new_bid

    # def _update_summary_sheet(self, wb, sheet_name, bidder_stats, min_row_start):
    #     """
    #     通用更新汇总表的方法
    #     :param wb: Workbook对象
    #     :param sheet_name: Sheet名称
    #     :param bidder_stats: 待更新的统计数据 {bidder_name: {project_count, total_amount}}
    #     :param min_row_start: 数据起始行号（即第一条数据所在的行）
    #     """
    #     if sheet_name not in wb.sheetnames:
    #         self.log(f"错误: 汇总文件中未找到 '{sheet_name}' Sheet")
    #         return 0, 0

    #     ws = wb[sheet_name]
        
    #     # 读取表头 (表头在第1行)
    #     header = [cell.value for cell in ws[1]]
    #     try:
    #         idx_name = header.index("公司名称")
    #         idx_count = header.index("中标总个数")
    #         idx_amount = header.index("中标总金额（万元）")
    #         idx_time = header.index("修改时间")
    #         idx_no = header.index("序号")
    #     except ValueError as e:
    #         self.log(f"错误: [{sheet_name}]表头缺失关键列 - {e}")
    #         return 0, 0

    #     # 建立现有行索引
    #     existing_rows = {}
    #     # start应与min_row一致，确保row_idx对应Excel实际行号
    #     for row_idx, row in enumerate(ws.iter_rows(min_row=min_row_start, values_only=True), start=min_row_start):
    #         raw_name = row[idx_name]
    #         if raw_name:
    #             existing_rows[raw_name] = row_idx

    #     # 样式定义
    #     font_style = Font(name="微软雅黑", size=8)
    #     align_style = Alignment(horizontal="center", vertical="center")
    #     border_style = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    #     updated_count = 0
    #     added_count = 0

    #     for bidder, stats in bidder_stats.items():
    #         if bidder in existing_rows:
    #             # 更新
    #             r_idx = existing_rows[bidder]
    #             ws.cell(row=r_idx, column=idx_count+1).value = stats["project_count"]
    #             ws.cell(row=r_idx, column=idx_amount+1).value = stats["total_amount"]
    #             ws.cell(row=r_idx, column=idx_time+1).value = f"{pd.Timestamp.now().strftime('%Y-%m-%d')}更新"
    #             updated_count += 1
    #             self.log(f"[{sheet_name}] 更新: {bidder}，项目数: {stats['project_count']}，总金额: {stats['total_amount']:.2f} 万元")
    #         else:
    #             # 新增
    #             new_row = [None] * len(header)
    #             new_row[idx_name] = bidder
    #             new_row[idx_count] = stats["project_count"]
    #             new_row[idx_amount] = stats["total_amount"]
    #             new_row[idx_time] = f"{pd.Timestamp.now().strftime('%Y-%m-%d')}新增"
    #             # 序号计算
    #             new_row[idx_no] = ws.max_row - (min_row_start - 1) + 1
                
    #             ws.append(new_row)
    #             added_count += 1
    #             self.log(f"[{sheet_name}] 新增: {bidder}，项目数: {stats['project_count']}，总金额: {stats['total_amount']:.2f} 万元")
                
    #             # 样式
    #             current_row = ws.max_row
    #             for col_idx in range(1, len(new_row) + 1):
    #                 cell = ws.cell(row=current_row, column=col_idx)
    #                 cell.font = font_style
    #                 cell.alignment = align_style
    #                 cell.border = border_style
        
    #     return updated_count, added_count

    def run_project_analysis_workflow(self):
        try:
            source_path = self.source_file_path.get()
            config_path = self.config_file_path.get()
            project_path = self.project_file_path.get()

            self.log(f"正在加载配置文件: {config_path}")
            
            # 加载关键词
            try:
                df_industry_kw = pd.read_excel(config_path, sheet_name="行业关键词")
                df_vendor_kw = pd.read_excel(config_path, sheet_name="厂商关键词")
            except Exception as e:
                self.log(f"错误: 读取配置文件失败 - {str(e)}")
                self.log("请确保配置文件包含 '行业关键词' 和 '厂商关键词' 两个Sheet")
                self.status_msg.set("配置读取失败")
                return

            # 构建关键词字典
            self.industry_map = {}
            for _, row in df_industry_kw.iterrows(): # iterrows()方法用于逐行遍历DataFrame，返回每一行的索引和数据（作为Series对象）。索引从0开始。
                ind = row['行业名称']
                kw = str(row['关键词']).strip()
                if ind not in self.industry_map:
                    self.industry_map[ind] = []
                if kw and kw != 'nan':
                    self.industry_map[ind].append(kw)

            self.vendor_map = {}
            for _, row in df_vendor_kw.iterrows():
                v_type = row['厂商类型']
                kw = str(row['关键词']).strip()
                if v_type not in self.vendor_map:
                    self.vendor_map[v_type] = []
                if kw and kw != 'nan':
                    self.vendor_map[v_type].append(kw)

            self.log("关键词加载完成。")
            self.log(f"正在读取源数据: {source_path}")

            # 1.根据文件名选择处理逻辑，将项目数据整理为统一格式
            filename = os.path.basename(source_path)
            df_new_bid = None
            
            self.status_msg.set("正在处理源数据...")
            if "ICT" in filename:
                df_new_bid = self._process_ict_data(source_path)
            elif "派单分析" in filename:
                df_new_bid = self._process_province_data(source_path)
            elif "数说123" in filename:
                df_new_bid = self._process_shushuo_data(source_path)
            else:
                self.log("警告: 文件名未包含已知来源标识(ICT/派单分析/数说123)，请检查文件名...")
                self.status_msg.set("未知源数据文件名")
                messagebox.showwarning("提示", "文件名未包含已知来源标识(ICT/派单分析/数说123)，请检查文件名后重试。")
                df_new_bid = None

            if df_new_bid is None or df_new_bid.empty:
                self.log("无新增项目信息，流程终止。")
                return
            
            self.log(f"源数据整理完成。正在检查重复并写入项目清单...")
            self.status_msg.set("正在更新[项目清单]...")

            # 2. 新项目检查去重后添加到项目汇总文件"项目清单"页（使用openpyxl）
            wb = openpyxl.load_workbook(project_path)
            if "项目清单" not in wb.sheetnames:
                self.log("错误: 汇总文件中未找到 '项目清单' Sheet")
                self.status_msg.set("汇总文件错误")
                return
            ws_bid = wb["项目清单"]
            
            # 获取现有项目数据用于查重、统计
            existing_bid_data = {} # 使用字典存储，key为中标单位，value为该单位的项目列表
            total_existing_count = 0 # 记录总行数，用于序号生成

            # 第一行是表头，从第二行开始读取
            # 需要确定关键列的索引（从1开始）
            header = [cell.value for cell in ws_bid[1]]
            try:
                idx_tenderer = header.index("招标单位")
                idx_bidder = header.index("中标单位")
                idx_amount = header.index("中标金额（万元）")
                idx_project = header.index("项目名称")
            except ValueError as e:
                self.log(f"错误: 汇总文件表头缺失关键列 - {e}")
                self.status_msg.set("汇总文件错误")
                return

            for row in ws_bid.iter_rows(min_row=2, values_only=True): # row是一个元组
                bidder_name = str(row[idx_bidder]).strip()
                if bidder_name not in existing_bid_data:
                    existing_bid_data[bidder_name] = []
                
                existing_bid_data[bidder_name].append({
                    "tenderer": str(row[idx_tenderer]).strip(),
                    "bidder": bidder_name,
                    "amount": float(row[idx_amount]) if row[idx_amount] else 0.0,
                    "project": str(row[idx_project]).strip()
                })
                total_existing_count += 1

            # 查重函数：项目名称相似度检查
            def check_project_name_duplicate(new_name, exist_name):
                if not new_name or pd.isna(new_name):
                    return False
                
                # 预处理函数：去除干扰词
                def clean_name(text):
                    text = str(text).strip()
                    # 去除常见的后缀和干扰词
                    keywords = ["招标", "中标", "成交", "结果", "公告", "公示", "项目", "采购", "关于", "的"]
                    for kw in keywords:
                        text = text.replace(kw, "")
                    # 去除标点和特殊字符
                    text = re.sub(r'[^\w\u4e00-\u9fa5]', '', text)
                    return text

                cleaned_new = clean_name(new_name)
                cleaned_exist = clean_name(exist_name)
                if len(cleaned_exist) < 4 or len(cleaned_new) < 4: # 太短的不查重，避免误判
                    return False
                    
                # 计算相似度：使用最长公共子串占比
                # 逻辑：如果短字符串的大部分内容出现在长字符串中，则视为重复
                matcher = difflib.SequenceMatcher(None, cleaned_new, cleaned_exist) # SequenceMatcher参数分别为：第一个参数是一个可选的函数，用于过滤掉不需要比较的字符，通常设置为None表示不进行过滤；第二个和第三个参数是要比较的两个字符串。
                match_size = sum(block.size for block in matcher.get_matching_blocks()) # get_matching_blocks()方法返回一个匹配块的列表，每个匹配块是一个包含三个属性的对象，分别是a、b和size，表示在第一个字符串和第二个字符串中匹配的起始位置以及匹配的长度。通过对所有匹配块的size属性求和，可以得到两个字符串之间的总匹配长度。
                
                shorter_len = min(len(cleaned_new), len(cleaned_exist))
                ratio = match_size / shorter_len
                
                # 阈值设定：0.85表示短字符串有85%以上的内容在长字符串中匹配
                if ratio > 0.85:
                    return True
                return False
            
            # 项目查重函数：综合判断
            def is_duplicate(new_row, existing_data):
                new_tenderer = str(new_row['招标单位']).strip() if not pd.isna(new_row['招标单位']) else ""
                new_bidder = str(new_row['中标单位']).strip() if not pd.isna(new_row['中标单位']) else ""
                new_name = str(new_row['项目名称']).strip() if not pd.isna(new_row['项目名称']) else ""
                try:
                    new_amount = float(new_row['中标金额（万元）'])
                except (ValueError, TypeError):
                    new_amount = 0.0
                
                # 优化：只在对应中标单位的项目列表中查找
                candidates = existing_data.get(new_bidder, [])
                
                # 查重：招标单位、中标单位、金额、项目名称
                for exist in candidates:
                    if (exist['tenderer'] == new_tenderer and 
                        exist['bidder'] == new_bidder and 
                        abs(exist['amount'] - new_amount) < 0.001 and
                        check_project_name_duplicate(new_name, exist['project'])):
                        return True
                    
                return False
            
            added_count = 0
            skipped_count = 0
            affected_bidders_set = set() # 记录新增项目里涉及到的中标单位，用于更新全量中标公司sheet

            # 遍历新项目数据并追加至项目清单Sheet
            project_no_init = total_existing_count + 1 # 用于序号递增
            for _, row in df_new_bid.iterrows(): # row是一个Series对象
                if is_duplicate(row, existing_bid_data):
                    self.log(f"项目：{row['项目名称']} 已存在，跳过。")
                    skipped_count += 1
                    continue
                
                row_no = project_no_init + added_count
                row['序号'] = row_no
                row['修改时间'] = f"{pd.Timestamp.now().strftime('%Y-%m-%d')}新增"
                # 转换为列表，准备追加
                row_values = [row[col] for col in self.target_columns]
                ws_bid.append(row_values)
                
                # 获取当前写入的行号
                current_row = ws_bid.max_row
                
                # 定义样式对象
                font_style = Font(name="微软雅黑", size=8)
                align_style = Alignment(horizontal="center", vertical="center")
                border_style = Border(
                    left=Side(style="thin"), 
                    right=Side(style="thin"), 
                    top=Side(style="thin"), 
                    bottom=Side(style="thin")
                )

                for col_idx in range(1, len(row_values) + 1):
                    cell = ws_bid.cell(row=current_row, column=col_idx)
                    cell.font = font_style
                    cell.alignment = align_style
                    cell.border = border_style

                added_count += 1
                # 更新existing_data以便后续查重（防止源数据内部重复）、统计
                affected_bidder = str(row['中标单位']).strip() if not pd.isna(row['中标单位']) else ""
                if affected_bidder not in existing_bid_data:
                    existing_bid_data[affected_bidder] = []
                    
                existing_bid_data[affected_bidder].append({
                    "tenderer": str(row['招标单位']).strip() if not pd.isna(row['招标单位']) else "",
                    "bidder": affected_bidder,
                    "amount": float(row['中标金额（万元）']) if not pd.isna(row['中标金额（万元）']) else 0.0,
                    "project": str(row['项目名称']).strip() if not pd.isna(row['项目名称']) else ""
                })
                
            self.log(f"项目信息处理完成。新增 {added_count} 条，跳过重复 {skipped_count} 条。")
            

            # 文件名加时间后缀另存
            new_project_file_path = re.sub(r'(\.xlsx|\.xls)$', f"_更新项目_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}\\1", project_path, flags=re.IGNORECASE) # \\1表示引用第一个捕获组
            try:
                wb.save(new_project_file_path)
            except Exception as e:
                self.log(f"保存项目汇总文件出错: {str(e)}")
                self.status_msg.set("保存项目汇总文件出错")
                messagebox.showerror("错误", f"保存项目汇总文件出错: {str(e)}")
                return
            self.log(f"处理完成。新增项目 {added_count} 条，跳过项目 {skipped_count} 条")
            self.status_msg.set("项目更新完成")
            messagebox.showinfo("成功", f"处理完成！\n新增项目: {added_count}\n跳过项目: {skipped_count}\n结果已保存至: {new_project_file_path}")

        except Exception as e:
            self.log(f"项目信息整理出错: {str(e)}")
            self.status_msg.set("项目信息整理出错")
            self.log(traceback.format_exc())

    def run_company_analysis_workflow(self):
        self.status_msg.set("正在加载公司汇总文件...")
        company_path = self.company_file_path.get()

        # 定义常量
        TARGET_YEARS = ["24年", "25年", "26年", "27年"] # 中标公司sheet列标签中统计的年份, 必须与文件保持一致
        TARGET_INDUSTRIES = [  # 中标公司sheet列标签中统计的行业，必须与文件保持一致
             "党政", "要客", "卫健", "住建应急", 
             "农业文宣", "教育", "金融交通", "工业"
            ]
        FILTER_AMOUNT = 400.0 # 筛选金额阈值，单位：万元

        # 生成年份统计列标签
        year_stat_col = []
        for year in TARGET_YEARS:
            year_stat_col.append("20" + year + "中标个数")
            year_stat_col.append("20" + year + "中标金额（万元）")
        # 生成行业统计列标签
        industry_stat_col = []
        for industry in TARGET_INDUSTRIES:
            industry_stat_col.append(industry + "中标个数")
            industry_stat_col.append(industry + "中标金额（万元）")
        # 生成中标公司sheet的列标签
        shared_company_stat_col = (
            ["序号", "公司名称", "设备商", "中标总个数", "中标总金额（万元）"] 
            + year_stat_col 
            + ["企业性质", "是否本地有办事处", "主要业务方向（重点）", "注册地", "登记机关", 
               "注册金额（万元）", "联系人", "联系方式", "重点企业经销商", "重点企业名称", "经销商等级"] 
            + industry_stat_col
            )
        
        all_company_stat_col = shared_company_stat_col+ ["修改时间"]
        filtered_company_stat_col = shared_company_stat_col + ["清单客户属地", "分配行业", "分配行业批次", "分配时间", "入库情况", "确认行业", "备注", "修改时间"]

        try:
            # 1. 读取数据并预处理
            self.log("正在读取公司汇总文件...")
            self.status_msg.set("正在读取公司汇总文件...")
            try: 
                df_all_project = pd.read_excel(company_path, sheet_name="项目清单")
                self.log("项目清单读取完成！")
                df_pre_all_company_stats = pd.read_excel(
                    company_path, sheet_name="全量中标公司", skiprows=2, header=None, 
                    usecols=range(len(all_company_stat_col)), names=all_company_stat_col
                    )
                self.log("全量中标公司读取完成！")
                self.log("全量中标公司页列标签设定如下，请确保与文件一致：")
                for idx, col in enumerate(all_company_stat_col, start=1):
                    self.log(f"{idx}. {col}")
                df_pre_filterd_company_stats = pd.read_excel(
                    company_path, sheet_name="筛后中标公司", skiprows=2, header=None, 
                    usecols=range(len(filtered_company_stat_col)), names=filtered_company_stat_col
                    )
                self.log("筛后中标公司读取完成！")
                self.log("筛后中标公司页列标签设定如下，请确保与文件一致：")
                for idx, col in enumerate(filtered_company_stat_col, start=1):
                    self.log(f"{idx}. {col}")
                wb = openpyxl.load_workbook(company_path)
            except Exception as e:
                self.log(f"错误：读取公司汇总文件失败 - {str(e)}")
                self.log("请确保公司汇总文件包含'项目清单', '全量中标公司', '筛后中标公司' Sheet")
                self.status_msg.set("公司汇总文件读取失败")
                return
            
            # 数据预处理
            # 清理空值，保证数值列数据类型正确
            count_cols = [col for col in shared_company_stat_col if "个数" in col]
            df_pre_all_company_stats[count_cols] = df_pre_all_company_stats[count_cols].apply(
                lambda x: pd.to_numeric(x, errors='coerce').fillna(0).astype(int)
                )
            df_pre_filterd_company_stats[count_cols] = df_pre_filterd_company_stats[count_cols].apply(
                lambda x: pd.to_numeric(x, errors='coerce').fillna(0).astype(int)
                )
            
            amount_cols = [col for col in shared_company_stat_col if "金额" in col]
            df_pre_all_company_stats[amount_cols] = df_pre_all_company_stats[amount_cols].apply(
                lambda x: pd.to_numeric(x, errors='coerce').fillna(0.0).astype('float64')
                )
            df_pre_filterd_company_stats[amount_cols] = df_pre_filterd_company_stats[amount_cols].apply(
                lambda x: pd.to_numeric(x, errors='coerce').fillna(0.0).astype('float64')
                )
            # # 一次性清洗所有类型的空格：半角、全角、制表符、换行符
            df_pre_all_company_stats["公司名称"] = (
                df_pre_all_company_stats["公司名称"]
                .astype(str)
                .str.strip()  # 先删首尾半角空格
                .str.replace(r"\s+", "", regex=True)  # 删所有空白字符（含全角/制表符/换行）
                )
            df_pre_filterd_company_stats["公司名称"] = (
                df_pre_filterd_company_stats["公司名称"]
                .astype(str)
                .str.strip()  # 先删首尾半角空格
                .str.replace(r"\s+", "", regex=True)  # 删所有空白字符（含全角/制表符/换行）
                )
            
            # 剔除中标单位为空或无效值的项目记录
            self.log("正在预处理项目数据...")
            invalid_values = ["", "无", "-", "/", "未知"]
            df_valid_project = df_all_project[~(df_all_project["中标单位"].isna() | df_all_project["中标单位"].isin(invalid_values))].copy()
            # 一次性清洗所有类型的空格：半角、全角、制表符、换行符
            df_valid_project["中标单位"] = (
                df_valid_project["中标单位"]
                .astype(str)
                .str.strip()  # 先删首尾半角空格,只要操作的是pandas的Series对象（而非单个 Python 原生字符串），调用任何字符串方法都必须通过.str访问器
                .str.replace(r"\s+", "", regex=True)  # 删所有空白字符（含全角/制表符/换行）
                )
            # 再次清除空字符串""
            df_valid_project = df_valid_project[df_valid_project["中标单位"] != ""].copy()
            self.log(f"原始项目记录 {len(df_all_project)} 条, 过滤无中标单位记录 {len(df_all_project) - len(df_valid_project)} 条")
            # 只统计非运营商的中标公司
            df_valid_project = df_valid_project[df_valid_project["中标厂商类型"] == "其他厂商"]
            self.log(f"过滤运营商中标项目后剩余 {len(df_valid_project)} 条")
            # 转换数据类型，处理空值
            df_valid_project["中标金额（万元）"] = pd.to_numeric(df_valid_project["中标金额（万元）"], errors='coerce').fillna(0.0) # coerce参数用于将无法转换的值设置为NaN，fillna(0)将NaN替换为0
            df_valid_project["中标年份"] = df_valid_project["中标年份"].fillna("未知")
            df_valid_project["项目所属行业"] = df_valid_project["项目所属行业"].fillna("未分类")
            
            # 2. 按中标单位分组统计核心指标
            self.status_msg.set("正在统计中标公司数据...")
            self.log("正在统计中标公司数据...")

            def calculate_company_metrics(group):
                metrics = {}
                # 总指标
                metrics["中标总个数"] = len(group)
                metrics["中标总金额（万元）"] = group["中标金额（万元）"].sum()

                # 各年份指标
                for year in TARGET_YEARS:
                    year_data = group[group["中标年份"] == year]
                    metrics["20" + year + "中标个数"] = len(year_data)
                    metrics["20" + year + "中标金额（万元）"] = year_data["中标金额（万元）"].sum()

                # 各行业指标
                for industry in TARGET_INDUSTRIES:
                    industry_data = group[group["项目所属行业"] == industry]
                    metrics[industry + "中标个数"] = len(industry_data)
                    metrics[industry + "中标金额（万元）"] = industry_data["中标金额（万元）"].sum()

                return pd.Series(metrics)
            df_new_all_company_stats = df_valid_project.groupby("中标单位").apply(calculate_company_metrics, include_groups=False).reset_index() # groupby返回一个DataFrameGroupBy对象，apply方法将自定义函数应用于每个分组，reset_index用于重置索引，使得“中标单位”成为DataFrame的一列而不是索引。
            # 将列名“中标单位”改为“公司名称”，与公司sheet保持一致，以便后续合并
            df_new_all_company_stats.rename(columns={"中标单位": "公司名称"}, inplace=True)
            self.log(f"共统计出 {len(df_new_all_company_stats)} 家中标公司。")

            # 3. 更新[中标公司]统计信息
            def merge_company_stats(df_pre_company_stats, df_new_company_stats, sheet_name):
                """
                用merge外连接实现：保留原表所有行（含重复公司）+ 原始顺序，新增公司追加在后
                核心：辅助列按「行」分配，锁定整表行顺序(含重复公司）
                :param old_company_df: 原有公司统计DataFrame（含重复公司，原始行顺序）
                :param new_bidder_stats: 新统计的公司数据 dict{公司名: {project_count: 数值, total_amount: 数值}}
                :return: 合并后的最终DataFrame（保留原表所有行+顺序，新增公司在后）
                """

                self.status_msg.set(f"正在合并 [{sheet_name}] 数据...")
                self.log(f"正在合并  [{sheet_name}] 数据...")
                # ========== 1. 原表预处理（优化：清洗公司名称+唯一辅助列） ==========
                # 深拷贝避免修改原数据，重置索引保证连续
                df_pre_company_stats = df_pre_company_stats.copy().reset_index(drop=True) #重置索引保证连续
                # 每行分配唯一_merge_sort_key
                df_pre_company_stats["_merge_sort_key"] = df_pre_company_stats.index
                # 记录原表总行数
                df_pre_total_row = len(df_pre_company_stats)
                self.log(f"[{sheet_name}页]有原始记录 {df_pre_total_row} 条")
        
                # ========== 2. 外连接合并，已有公司更新，新增公司追加 ==========
                df_merged_company_stats  = pd.merge(
                    df_pre_company_stats, 
                    df_new_company_stats,
                    on="公司名称", how="outer",
                    suffixes=("", "_new")
                    ) # 同名列，原有数据列名不变，新数据列名加后缀_new,会产生两列，不会覆盖

                # ========== 3. 标记列变化，更新修改时间 ==========
                new_cols = [col for col in df_merged_company_stats.columns if col.endswith("_new")]
                changed_cols = []
                for new_col in new_cols:
                    orig_col = new_col.replace("_new", "")
                    changed_col = f"{orig_col}_changed"
                    changed_cols.append(changed_col)
                    # 判断列是否有变化，浮点数列单独处理
                    col_dtype = df_merged_company_stats[orig_col].dtype
                    if np.issubdtype(col_dtype, np.floating):
                        df_merged_company_stats[changed_col] = ~np.isclose(
                            df_merged_company_stats[orig_col],
                            df_merged_company_stats[new_col],
                            atol=1e-6, # 绝对误差≤0.000001即判定为相等（适配万元级金额，足够精准）
                            equal_nan=True
                            )
                    else: # 整数、字符等其它类型，无精度问题
                        df_merged_company_stats[changed_col] = ~df_merged_company_stats[orig_col].eq(df_merged_company_stats[new_col])

                # 有变化的行、更新修改时间
                df_merged_company_stats.loc[df_merged_company_stats[changed_cols].any(axis=1), "修改时间"] = pd.Timestamp.now().strftime("%Y-%m-%d") + "修改"
                # _merge_sort_key列为无效值，将修改时间设置为新增
                df_merged_company_stats.loc[df_merged_company_stats["_merge_sort_key"].isna(), "修改时间"] = pd.Timestamp.now().strftime("%Y-%m-%d") + "新增"
                # 新增行_sort_key列填充为大于原始最大值的序号，确保排序在最后
                df_merged_company_stats.loc[df_merged_company_stats["_merge_sort_key"].isna(), "_merge_sort_key"] = df_pre_total_row + 100

                # ========== 3. 新列中的无效值，使用原始列填充，之后用新列替换原始列 ==========
                for new_col in new_cols:
                    orig_col = new_col.replace("_new", "")
                    df_merged_company_stats[orig_col] = df_merged_company_stats[new_col].fillna(df_merged_company_stats[orig_col])
                # 删除新列
                df_merged_company_stats.drop(columns=new_cols, inplace=True)
                
                # ========= 4. 恢复原表顺序、新增行在后；重新生成序号 ==========
                # 按_merge_sort_key排序
                df_merged_company_stats.sort_values(by="_merge_sort_key", inplace=True, ignore_index=True) # ignore_index=True参数用于在排序后重置索引，确保索引连续且从0开始
                # 序号重新生成
                df_merged_company_stats["序号"] = range(1, len(df_merged_company_stats) + 1)

                # 删除辅助列
                df_merged_company_stats.drop(columns=["_merge_sort_key"] + changed_cols, inplace=True)

                # 日志记录新增公司
                if len(df_merged_company_stats) > df_pre_total_row:
                    # itertuples(index=False)：不返回索引，只返回行数据；命名元组的属性名对应列名（特殊字符会自动替换）
                    for row_tuple in df_merged_company_stats[df_pre_total_row:].itertuples(index=False):
                        company_name = getattr(row_tuple, "公司名称", None)
                        bid_count = getattr(row_tuple, "中标总个数", 0)
                        bid_amount = getattr(row_tuple, "中标总金额（万元）", 0.0)
                        self.log(f"[{sheet_name}] 新增公司: {company_name}, 中标总个数: {bid_count}, 中标总金额（万元）: {bid_amount:.2f} 万元")
                return df_merged_company_stats

            # 4. 写回Excel文件，使用openpyxl设置格式
            def write_company_stats_to_sheet(wb, sheet_name, df_stats, start_row):
                self.status_msg.set(f"正在写入[{sheet_name}] sheet...")
                self.log(f"正在写入[{sheet_name}] sheet...")

                ws = wb[sheet_name]

                # 定义样式
                font_style = Font(name="微软雅黑", size=8)
                align_style = Alignment(horizontal="center", vertical="center")
                border_style = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )

                # 清空旧数据（保留表头）
                ws.delete_rows(start_row, ws.max_row - start_row + 1) # 两个参数分别是起始行号和删除的行数

                # 写入新数据并设置样式
                for r_idx, row in enumerate(dataframe_to_rows(df_stats, index=False, header=False), start=start_row): # index=False表示不写入索引列，header=False表示不写入列名
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws.cell(r_idx, c_idx, value=value)
                        cell.font = font_style
                        cell.alignment = align_style
                        cell.border = border_style

                self.log(f"[{sheet_name}] sheet写入完成")

            # 更新并写入[全量中标公司]统计信息
            df_merged_all_company_stats = merge_company_stats(df_pre_all_company_stats, df_new_all_company_stats, "全量中标公司")
            write_company_stats_to_sheet(wb, "全量中标公司", df_merged_all_company_stats, start_row=3)
            # 更新并写入[筛后中标公司]统计信息
            df_new_filterd_company_stats = df_new_all_company_stats[df_new_all_company_stats["中标总金额（万元）"] > FILTER_AMOUNT]
            df_merged_filterd_company_stats = merge_company_stats(df_pre_filterd_company_stats, df_new_filterd_company_stats, "筛后中标公司")
            write_company_stats_to_sheet(wb, "筛后中标公司", df_merged_filterd_company_stats, start_row=3)

            # 保存文件
            new_company_file_path = re.sub(r'(\.xlsx|\.xls)$', f"_更新公司_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}\\1", company_path, flags=re.IGNORECASE) # \\1表示引用第一个捕获组
            try:
                wb.save(new_company_file_path)
            except Exception as e:
                self.log(f"保存公司统计文件出错: {str(e)}")
                self.status_msg.set("保存公司统计文件出错")
                messagebox.showerror("错误", f"保存公司统计文件出错: {str(e)}")
                return
            self.status_msg.set("公司统计分析完成")
            self.log("公司统计分析完成。")
            messagebox.showinfo("成功", f"公司统计分析完成！\n全量中标公司新增 {len(df_merged_all_company_stats) - len(df_pre_all_company_stats)} 家\n筛后中标公司新增 {len(df_merged_filterd_company_stats) - len(df_pre_filterd_company_stats)} 家\n结果已保存至: {new_company_file_path}")
            
        except Exception as e:
            self.log(f"错误：公司统计分析出错 - {str(e)}")
            self.status_msg.set("公司统计分析出错")
            self.log(traceback.format_exc())

if __name__ == "__main__":
    root = tk.Tk() # 创建主窗口
    app = BidAnalysisApp(root)
    root.mainloop() # 进入Tkinter事件循环
