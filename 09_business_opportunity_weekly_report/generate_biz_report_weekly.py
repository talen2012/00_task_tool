import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk # ttk是tkinter的一个子模块，提供了更现代化的组件，如Combobox、Treeview等
from tkcalendar import DateEntry # 日历选择框
import pandas as pd
from pandas.tseries.offsets import MonthEnd
import openpyxl
import os
import re
import yaml
from typing import Dict
from datetime import datetime
import logging
import traceback

# @Time     : 2026/04/13/16:00
# @Author   : talen
# @File     : generate_biz_report_weekly.py

class NameMappingError(Exception):
    """单元/行业名称映射失败的自定义异常"""
class BizReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("商机报告生成工具")
        self.root.geometry("800x600")
        # ============================= 用户输入变量 =============================
        self.biz_above_million_filepath = tk.StringVar()
        self.biz_below_million_filepath = tk.StringVar()
        self.contract_status_filepath = tk.StringVar()
        self.biz_code_filepath = tk.StringVar()
        self.status_msg = tk.StringVar(value="等待用户输入")

        self.start_date = tk.StringVar()
        self.end_date = tk.StringVar()
        # 供统计月下拉选择
        self.stat_year = tk.StringVar(value=pd.Timestamp.now().strftime("%Y"))
        self.stat_month = tk.StringVar(value=pd.Timestamp.now().strftime("%m"))
        self.years_for_select = [str(y) for y in range(2020, 2076)]
        self.months_for_select = [f"{m:02d}" for m in range(1, 13)]
        # 统计口径，每种对应一个统计结果文件，允许同时选择

        self.strict_mode = tk.BooleanVar() # 签约情况中的合同，必须有对应商机编码，并且该商机编码存在于商机清单中，否则不计入统计
        self.loose_mode = tk.BooleanVar(value=True) # 不考虑合同与商机编码的对应关系，直接统计签约情况中的合同数据

        self.generate_btn = None # 初始化生成报告按钮，后续动态控制状态

        # ============ reference目录下的文件路径是写死的，不要移动 ==================
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.config_filepath = os.path.join(self.base_dir, "reference", "config.yaml")
        self.unit_industry_map_filepath = os.path.join(self.base_dir, "reference", "单元_行业名称映射表.xlsx")
        self._report_template_filepath = os.path.join(self.base_dir, "reference", "通报模板.xlsx")

        # =============================== 创建UI ==================================
        self._create_ui()

        # ============== 从配置文件读取目标文件名、sheet名及必须列等 ===================
        # 百万以上
        self.above_million_filename_keyword = None
        self.above_million_sheet = None
        self.above_million_required_cols = None
        # 百万以下
        self.below_million_filename_keyword = None
        self.below_million_sheet = None
        self.below_million_required_cols = None
        # 签约情况
        self.contract_status_filename_keyword = None
        self.contract_status_sheet = None
        self.contract_status_required_cols = None
        # 全量商机编码清单
        self.biz_code_filename_keyword = None
        self.biz_code_sheet = None
        self.biz_code_required_cols = None
        # 通报模板中的行业、单元名称，保留顺序
        self.unit_order = None
        self.industry_order = None
        # 明细列标签，保留顺序
        self.col_order = None

        # 防抖定时器，由于延迟校验输入
        self._check_timer = None

        # ================================= 初始化日志 ============================
        logging.basicConfig(
            level=logging.INFO, # DEBUG < INFO < WARNING < ERROR < CRITICAL, 日志级别依次递增，设置为INFO表示记录INFO及以上级别的日志
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(os.path.join(self.base_dir, 'biz_report_log.log'), encoding='utf-8'),
                logging.StreamHandler() # 同时输出到控制台，方便调试
            ]
        )
        self.logger = logging.getLogger(__name__) # 创建一个logger对象，__name__表示当前模块的名字，使用模块名作为logger名字是常见的做法，可以区分不同模块的日志
        # 如果__name__ == "__main__"，表示直接运行这个脚本，此时logger名字就是"__main__"；如果这个脚本被其他模块导入，logger名字就是模块的名字，这样可以区分日志来源
        
        # ============================= 配置文件加载及状态 =========================
        self.config_ok = False
        self._load_config()
        
        # =============== 自动查找程序目录下可能的商机清单及签约情况文件 ==============
        self._auto_find_files()

        # ===================== 绑定输入变化监听事件（实时校验输入）==================
        self._bind_input_trace()
        
        # ========================= 初始检查用户输入完整度 ==========================
        self._check_input_completeness()

    def _auto_find_files(self):
        if not self.config_ok:
            return
        for file_name in os.listdir(self.base_dir):  # listdir仅第一层，不递归，返回字符串列表，os.walk可遍历
            # 只匹配excel文件
            if file_name.startswith(("~$")): continue # 跳过隐藏文件
            if file_name.lower().endswith(('.xlsx', '.xls')):
                if self.above_million_filename_keyword in file_name:
                    self.biz_above_million_filepath.set(os.path.join(self.base_dir, file_name))
                    self._log(f"自动匹配到百万以上-商机文件：{file_name}")
                if self.below_million_filename_keyword in file_name:
                    self.biz_below_million_filepath.set(os.path.join(self.base_dir, file_name))
                    self._log(f"自动匹配到百万以下商机文件：{file_name}")
                if self.contract_status_filename_keyword in file_name:
                    self.contract_status_filepath.set(os.path.join(self.base_dir, file_name))
                    self._log(f"自动匹配到签约情况文件：{file_name}")
                if self.biz_code_filename_keyword in file_name:
                    self.biz_code_filepath.set(os.path.join(self.base_dir, file_name))
                    self._log(f"自动匹配到商机编码清单文件：{file_name}")

    def _create_ui(self):
        # ============================= 文件选择区域 =============================
        frame_files = tk.LabelFrame(self.root, text="文件选择", padx=10, pady=10) # 框架内部的内边距
        frame_files.pack(fill="x", padx=10, pady=5) # 框架和父窗口/其它组件的边距
        # 1. 百万以上商机清单
        tk.Label(frame_files, text="百万以上商机清单:").grid(row=0, column=0, sticky='w')
        tk.Entry(frame_files, textvariable=self.biz_above_million_filepath, width=80).grid(row=0, column=1, padx=5)
        tk.Button(frame_files, text="浏览...", command=lambda: self._select_file(self.biz_above_million_filepath)).grid(row=0, column=2)
        # 2. 百万以下商机清单
        tk.Label(frame_files, text='百万以下商机清单:').grid(row=1, column=0, sticky='w')
        tk.Entry(frame_files, textvariable=self.biz_below_million_filepath, width=80).grid(row=1, column=1, padx=5)
        tk.Button(frame_files, text='浏览...', command=lambda: self._select_file(self.biz_below_million_filepath)).grid(row=1, column=2)
        # 3. 签约情况表
        tk.Label(frame_files, text='商机签约情况表:').grid(row=2, column=0, sticky='w')
        tk.Entry(frame_files, textvariable=self.contract_status_filepath, width=80).grid(row=2, column=1, padx=5)
        tk.Button(frame_files, text='浏览...', command=lambda: self._select_file(self.contract_status_filepath)).grid(row=2, column=2)
        # 4.商机编码清单
        tk.Label(frame_files, text='商机编码全量清单:').grid(row=3, column=0, sticky='w')
        tk.Entry(frame_files, textvariable=self.biz_code_filepath, width=80).grid(row=3, column=1, padx=5)
        tk.Button(frame_files, text='浏览...', command=lambda: self._select_file(self.biz_code_filepath)).grid(row=3, column=2)
     
        # ================================= 设置 =================================
        frame_settings = tk.LabelFrame(self.root, text='设置', padx=5, pady=5)
        frame_settings.pack(fill='x', padx=10, pady=5)
        # 1. 日期选择
        frame_settings_date = tk.LabelFrame(frame_settings, text='日期选择', padx=10, pady=10)
        frame_settings_date.pack(side='left', padx=10, pady=5,anchor='n')
        tk.Label(frame_settings_date, text="开始日期:").grid(row=0, column=0, sticky='w')
        DateEntry(frame_settings_date, textvariable=self.start_date, date_pattern="yyyy-mm-dd", width=14).grid(row=0, column=1, padx=5)
        tk.Label(frame_settings_date, text='结束日期:').grid(row=0, column=2, padx=5)
        DateEntry(frame_settings_date, textvariable=self.end_date, date_pattern="yyyy-mm-dd", width=14).grid(row=0, column=3, padx=5)
        tk.Label(frame_settings_date, text="统计月:").grid(row=0, column=4, padx=5)
        ttk.Combobox(frame_settings_date, textvariable=self.stat_year, values=self.years_for_select, width=6).grid(row=0, column=5)
        ttk.Combobox(frame_settings_date, textvariable=self.stat_month, values=self.months_for_select, width=3).grid(row=0, column=6)

        # 2. 统计口径
        frame_settings_caliber = tk.LabelFrame(frame_settings, text="统计口径", padx=10, pady=10)
        frame_settings_caliber.pack(side='left', padx=10, pady=5,anchor='n')
        ttk.Checkbutton(frame_settings_caliber, text="宽口径", variable=self.loose_mode).grid(row=0, column=0, padx=5)
        ttk.Checkbutton(frame_settings_caliber, text='窄口径', variable=self.strict_mode).grid(row=0, column=1, padx=5)

        # =============================== 操作区域 ===============================
        frame_actions = tk.LabelFrame(self.root, padx=10, pady=10)
        frame_actions.pack(fill='x', side='top', padx=10, pady=5)
        # 初始化生成报告按钮（默认置灰）
        self.generate_btn = tk.Button(
            frame_actions,
            text="生成报告",
            command=self.run_biz_analysis_workflow,
            bg="#cccccc",
            fg="#666666",
            font=('Microsoft Yahei', 10, 'bold'),
            state=tk.DISABLED
            )
        self.generate_btn.pack(side='left', padx=5)
        
        tk.Label(frame_actions, textvariable=self.status_msg, fg='blue').pack(side='left', padx=50)

        # =============================== 日志区域 ===============================
        frame_log = tk.LabelFrame(self.root, text='运行日志', padx=10, pady=10)
        frame_log.pack(fill='both', expand=True, padx=10, pady=5)
        self.log_text = scrolledtext.ScrolledText(frame_log, height=20)
        self.log_text.pack(fill="both", expand=True)

    def _bind_input_trace(self):
        """绑定所有输入变量的变化监听（实时校验）"""
        # 监听文件路径的变化
        self.biz_above_million_filepath.trace_add("write", lambda *args: self._delayed_check())
        self.biz_below_million_filepath.trace_add("write", lambda *args: self._delayed_check())
        self.contract_status_filepath.trace_add("write", lambda *args: self._delayed_check())
        self.biz_code_filepath.trace_add("write", lambda *args: self._delayed_check())
        # 监听日期、统计月变化
        self.start_date.trace_add("write", lambda *args: self._delayed_check())
        self.end_date.trace_add("write", lambda *args: self._delayed_check())
        self.stat_year.trace_add("write", lambda *args: self._delayed_check())
        self.stat_month.trace_add("write", lambda *args: self._delayed_check())
        # 监听统计口径变化
        self.strict_mode.trace_add("write", lambda *args: self._delayed_check())
        self.loose_mode.trace_add("write", lambda *args: self._delayed_check())

    def _delayed_check(self, delay_ms=500):
        """
        输入防抖延迟校验：停止操作500ms后，再执行输入校验，避免频繁触发
        """
        # 如果上一个定时器还在，取消它
        if self._check_timer:
            self.root.after_cancel(self._check_timer)
        # 重新设置定时器，delay_ms毫秒后执行输入校验
        self._check_timer = self.root.after(delay_ms, self._check_input_completeness)

    def _check_input_completeness(self):
        """分类校验，失败弹窗+直接返回，更新状态和按钮样式"""
        def wait_input():
            self.generate_btn.config(bg="#cccccc", fg="#666666", state=tk.DISABLED)
            self._update_status("等待用户输入")
            self._log("等待用户输入...")
        # ============================= 文件路径校验 =============================
        # 1. 路径非空
        required_filepath_inputs = [
            self.biz_above_million_filepath.get().strip(),
            self.biz_below_million_filepath.get().strip(),
            self.contract_status_filepath.get().strip(),
            self.biz_code_filepath.get().strip()
        ]
        if not all(required_filepath_inputs):
            wait_input()
            return
        # 2. 路径有效
        file_var_map = [
            (self.biz_above_million_filepath, "百万以上商机文件"),
            (self.biz_below_million_filepath, "百万以下商机文件"),
            (self.contract_status_filepath, "签约情况文件"),
            (self.biz_code_filepath, "商机编码清单文件")
        ]
        for var, name in file_var_map:
            path = var.get().strip()
            if not os.path.exists(path):
                self.generate_btn.config(bg="#cccccc", fg="#666666", state=tk.DISABLED)
                self._update_status("文件路径有误")
                self._log(f'【{name}】不存在或为空！', 'error')
                messagebox.showerror("路径无效", f"【{name}】不存在或为空！")
                return
                
        # =============================== 日期校验 ===============================
        # 1. 日期非空且有效
        start = self.start_date.get().strip()
        end = self.end_date.get().strip()
        if start and end:
            try:
                start = pd.to_datetime(start)
                end = pd.to_datetime(end)
                if start > end:
                    self.generate_btn.config(bg="#cccccc", fg="#666666", state=tk.DISABLED)
                    self._log("开始日期不能晚于结束日期！", 'error')
                    self._update_status("日期输入有误")
                    messagebox.showerror("日期无效", "开始日期不能晚于结束日期！")
                    return
            except:
                self.generate_btn.config(bg="#cccccc", fg="#666666", state=tk.DISABLED)
                self._log("日期格式有误！", 'error')
                self._update_status("日期输入有误")
                messagebox.showerror("日期无效", "日期格式不正确")
                return
        else:
            wait_input()
            return

        # =============================== 月份校验 ===============================
        year_month_ok = self.stat_year.get().strip() and self.stat_month.get().strip()
        
        # =============================== 口径校验 ===============================
        mode_ok = self.loose_mode.get() or self.strict_mode.get()
        # 还有配置文件是否加载OK
        if year_month_ok and mode_ok and self.config_ok:
            self._update_status("就绪")
            self._log("所有输入已填写且校验通过")
            # 更新按钮样式（可点击）
            self.generate_btn.config(
                bg="#4CAF50",
                fg="white",
                state=tk.NORMAL  # 启用按钮
            )
        else:
            wait_input()
    def _update_status(self, message):
        self.status_msg.set(message)
        self.root.update()
        self.root.update_idletasks()

    def _log(self, message, level: str = 'info'):
        """
        统一日志格式按照logging模块：
        2026-04-29 09:26:02,039 - INFO - 自动匹配到文件
        同时输出到GUI、文件、控制台
        """
        timestamp = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S,%f')[:-3] # now()获取当前系统时间（精确到微秒），比如2026-04-29 09:26:02.039123
        level_upper = level.upper()
        log_line = f'{timestamp} - {level_upper} - {message}'
        self.log_text.insert(tk.END, log_line + '\n')
        self.log_text.see(tk.END)
        self.root.update() # 强制处理所有UI事件
        self.root.update_idletasks() # 强制绘制

        # 文件/控制台输出
        if level == 'warning':
            self.logger.warning(message)
        elif level == 'error':
            self.logger.error(message)
        else:
            self.logger.info(message)
    
    def _select_file(self, filepath_var: tk.StringVar):
        filename = filedialog.askopenfilename(
            initialdir=self.base_dir,
            filetypes=[('Excel file', '*.xlsx;*.xls')]
        )
        if filename:
            filepath_var.set(filename)

    def _load_config(self): 
        """
        加载配置文件
        """ 
        self._log("正在加载配置文件...")
        # ============================= 加载config文件 =============================
        self.config: Dict = None
        config_load_success = True
        try:
            # 1. 检查文件是否存在
            if not os.path.exists(self.config_filepath):
                raise FileNotFoundError(f"配置文件不存在！请检查路径：{self.config_filepath}")
            # 2. 读取并解析yaml文件
            with open(self.config_filepath, 'r', encoding='utf-8') as f:
                self.config = yaml.safe_load(f)
            # 3. 校验yaml是否为空
            if not self.config:
                raise ValueError(f"配置文件内容为空, 请检查：{self.config_filepath}")
        except PermissionError as pe:
            self._log(f"加载配置失败：\n无权限读取配置文件: {self.config_filepath}\n{traceback.format_exc()}", 'error')
            messagebox.showerror('错误', '加载配置失败')
            config_load_success = False
        except ValueError as ve:
            self._log(f"加载配置失败：\n{traceback.format_exc()}", 'error')
            messagebox.showerror('错误', '加载配置失败')
            config_load_success = False
        except FileNotFoundError as fnfe:
            self._log(f"加载配置失败：\n{traceback.format_exc()}", 'error')
            messagebox.showerror('错误', '加载配置失败')
            config_load_success = False
        except Exception as e:
            self._log(f"加载配置失败：\n{traceback.format_exc()}", 'error')
            messagebox.showerror('错误', '加载配置失败')
            config_load_success = False

        if config_load_success ==False:
            self.config_ok = False
            return

        # ===================== 读取配置（自动校验字段是否存在） =====================
        try:
            # 百万以上
            self.above_million_filename_keyword = self.config["above_million"]["workbook_name"]
            self.above_million_sheet = self.config["above_million"]["sheet_name"]
            self.above_million_required_cols = self.config["above_million"]["required_cols"]
            # 百万以下
            self.below_million_filename_keyword = self.config["below_million"]["workbook_name"]
            self.below_million_sheet = self.config["below_million"]["sheet_name"]
            self.below_million_required_cols = self.config["below_million"]["required_cols"]
            # 签约情况
            self.contract_status_filename_keyword = self.config["contract_status"]["workbook_name"]
            self.contract_status_sheet = self.config["contract_status"]["sheet_name"]
            self.contract_status_required_cols = self.config["contract_status"]["required_cols"]
            # 全量商机编码清单
            self.biz_code_filename_keyword = self.config['biz_code']['workbook_name']
            self.biz_code_sheet = self.config['biz_code']['sheet_name']
            self.biz_code_required_cols = self.config['biz_code']['required_cols']
            # 合同排除关键词
            # self.contract_exclude_keywords = self.config["contract_exclude_keywords"]
            # 通报模板中的单元、行业名称，固定顺序
            self.unit_order = self.config['unit_order']
            self.industry_order = self.config['industry_order']
            self.col_order = self.config['col_order']
        # 捕获：配置文件少写了字段（比如漏了 above_million）
        except KeyError as ke:
            self._log(f"加载配置失败：\n配置文件缺少关键字段: {traceback.format_exc()}", 'error')
            messagebox.showerror('错误', '加载配置失败')
            config_load_success = False
        except Exception as e:
            self._log(f"加载配置失败：\n{traceback.format_exc()}", 'error')
            messagebox.showerror('错误', '加载配置失败')
            config_load_success = False
        
        if config_load_success == True:
            self._log("成功加载配置文件！")
            self.config_ok = True
        
    def _open_source_table(self):
        """使用pandas打开底表文件, 并检查是否存在指定列"""
        # 要加载的文件列表
        self._update_status("读取文件")
        file_list = [
            (self.unit_industry_map_filepath, "单元", ["清单底表", "通报模板"], "单元名称映射文件"),
            (self.unit_industry_map_filepath, "行业", ["清单底表", "通报模板"], "行业名称映射文件"),
            (self.biz_above_million_filepath.get().strip(), self.above_million_sheet, self.above_million_required_cols, "百万以上商机文件"),
            (self.biz_below_million_filepath.get().strip(), self.below_million_sheet, self.below_million_required_cols, "百万以下商机文件"),
            (self.contract_status_filepath.get().strip(), self.contract_status_sheet, self.contract_status_required_cols, "签约情况文件"),
            (self.biz_code_filepath.get().strip(), self.biz_code_sheet, self.biz_code_required_cols, "全量商机编码文件")
        ]
        # 逐个加载
        def _check_na(df, col):
            return ((df[col].isna()) | (df[col] == '')).sum()
        df_list = []
        for filepath, sheetname, required_cols, display_name in file_list:
            self._log(f"正在加载{display_name}...")
            try:
                df = pd.read_excel(filepath, sheet_name=sheetname, header=1 if "映射文件" in display_name else 0, usecols=required_cols)
                df_list.append(df)
                self._log(f"成功加载{display_name}！共 {len(df)} 行")
                # 提醒关键列数据缺失
                if "百万" in display_name or "签约" in display_name:
                    self._log(f"{display_name}缺失：单元{_check_na(df, required_cols[1])}个，行业{_check_na(df, required_cols[2])}个，日期{_check_na(df, required_cols[3])}个", "warning")
            except ValueError as ve: # usecols列缺失，捕获ValueError
                self._log(f"{display_name}缺少必需列：\n{traceback.format_exc()}", 'error')
                self._update_status("读取文件失败")
                messagebox.showerror('错误', f'加载{display_name}失败')
                return None
            except Exception as e:
                self._log(f"{display_name}格式不正确：\n{traceback.format_exc()}", 'error')
                messagebox.showerror('错误', f'加载{display_name}失败')
                return None
        # 单独提醒全量商机编码文件关键列数据缺失
        self._log(f"全量商机编码文件缺失：日期{_check_na(df_list[5], self.biz_code_required_cols[1])}个", "warning")
        return df_list

    def _summarize_by_unit_and_industry(self, df_unit_map, df_industry_map, df_biz_above_million, df_biz_below_million, df_contract_status, df_biz_code): 
        """
        根据100万以上/以下商机清单，及合同签约情况生成每周商机进展报告
        @param df_biz_above_million: 100万以上商机清单
        @param df_biz_below_million: 100万以下商机清单
        @param df_contract_status: 合同实际签约情况
        @param df_unit_map: 单元名称映射表
        @param df_industry_map: 行业名称映射表
        @param df_biz_code: 全量商机编码及预计签约日期, 用于窄口径校验
        """
        # =============================== 数据预处理 ===============================
        self._update_status("数据预处理")
        self._log("开始进行数据预处理...")
        # # 1.删去签约情况表中，"比对"列内容包含"集成""陕数""省公司"的合同
        # contract_exclude_pattern = "|".join(self.contract_exclude_keywords)
        # df_contract_status_required = df_contract_status_required[
        #     ~df_contract_status_required[self.above_million_required_cols[0]].str.contains(contract_exclude_pattern, na=False) #na视为不包含关键词
        #     ].copy()
        # self._log(f"已清理\"{self.above_million_required_cols[0]}\"列含 {self.contract_exclude_keywords} 的签约数据，剩余 {len(df_contract_status_required)} 个！")
        
        # 1. 根据配置文件确定必需列的名称、统一列名
        # 百万以上
        above_million_bizcode_col_name, above_million_unit_col_name, above_million_industry_col_name, above_million_date_col_name, above_million_money_col_name, above_million_target_col_name, above_million_biz_name = self.above_million_required_cols
        # 百万以下
        below_million_bizcode_col_name, below_million_unit_col_name, below_million_industry_col_name, below_million_date_col_name, below_million_money_col_name, below_million_biz_name = self.below_million_required_cols
        # 签约情况
        contract_status_bizcode_col_name, contract_status_unit_col_name, contract_status_industry_col_name, contract_status_date_col_name, contract_status_money_col_name = self.contract_status_required_cols
        # 商机编码清单
        biz_code_bizcode_col_name, biz_code_date_col_name = self.biz_code_required_cols
        # 统一列名，方便后续合并和统计
        df_above = df_biz_above_million.rename(columns={
            above_million_bizcode_col_name: "biz_code",
            above_million_unit_col_name: "unit",
            above_million_industry_col_name: "industry",
            above_million_date_col_name: "date",
            above_million_money_col_name: "money",
            above_million_target_col_name: "target",
            above_million_biz_name: "name"}
            )
        df_above['above_flag'] = True # 标记百万以上商机，后续统计时可以区分百万以上/以下
        df_below = df_biz_below_million.rename(columns={
            below_million_bizcode_col_name: "biz_code",
            below_million_unit_col_name: "unit",
            below_million_industry_col_name: "industry",
            below_million_date_col_name: "date",
            below_million_money_col_name: "money",
            below_million_biz_name: "name"}
            )
        df_below['target'] = ""
        df_below['above_flag'] = False # 标记百万以下商机
        df_contract = df_contract_status.rename(columns={
            contract_status_bizcode_col_name: "biz_code",
            contract_status_unit_col_name: "unit",
            contract_status_industry_col_name: "industry",
            contract_status_date_col_name: "sign_date",
            contract_status_money_col_name: "money"
        })
        df_bizcode = df_biz_code.rename(columns={
            biz_code_bizcode_col_name: "biz_code",
            biz_code_date_col_name: "date"
        })
        
        # 2. 确保数据类型正确
        # # 日期列
        df_above['date'] = pd.to_datetime(
            df_above['date'],
            format="%Y%m", # 底表格式是202604
            errors="coerce" # 无法转换时填充为空，不报错
        )#.dt.strftime("%Y/%m/%d") # 自动补1号，例如20260401
        df_below['date'] = pd.to_datetime(
            df_below['date'],
            errors="coerce"
        )
        df_contract['sign_date'] = pd.to_datetime(
            df_contract['sign_date'],
            errors="coerce"
        )
        df_bizcode['date'] = pd.to_datetime(
            df_bizcode['date'],
            errors='coerce'
        )
        # # 数值列
        df_above['money'] = pd.to_numeric(
            df_above['money'],
            errors="coerce" # 无法转换时填充为空，不报错
        ).fillna(0).astype('float64')
        df_below['money'] = pd.to_numeric(
            df_below['money'],
            errors="coerce"
        ).fillna(0).astype('float64')
        df_contract['money'] = pd.to_numeric(
            df_contract['money'],
            errors="coerce"
        ).fillna(0).astype('float64')
        # # 将签约情况表中的"省口径签约额"列单位从"元"改为"万元"
        df_contract['money'] = df_contract['money'] / 10000
        
        # 3. 构建单元名称映射表、行业名称映射表，可能存在多对一映射关系
        ## 注意不要用dict，因为清单底表列存在"高新","南高新"这种相互包含的关键字
        ## 并且本来就是要逐个遍历，所以即使使用dict也不会提高效率
        ## 因此使用list保留顺序, 并根据关键字长度排个序，长的关键字在前边，避免"南高新"被"高新"给匹配走
        unit_map_list = df_unit_map[["清单底表", "通报模板"]].values.tolist() #values是将DataFrame转换成二维numpy数组，tolist()再将其转换成列表
        unit_map_list.sort(key=lambda x: len(x[0]), reverse=True)
        industry_map_list = df_industry_map[["清单底表", "通报模板"]].values.tolist()
        industry_map_list.sort(key=lambda x: len(x[0]), reverse=True)

        # 4. 根据单元_行业名称映射表，将底表的单元、行业两列映射成与模板中一致
        # 注意：非绝对匹配，只要底表名称包含"清单底表"列的某个键，就可以映射到对应的"通报模板"名称，不成功会标记为"unmatched"
        def map_fuzzy_names(data_series, map_list, type: str):
            """
            模糊名称匹配，原始值包含关键词→替换为目标值
            """
            unmatched_names = [] # 收集所有未匹配的名称
            def map_func(x):
                if pd.isna(x):
                    return "unmatched"
                x_str = str(x).strip()
                for key, target_name in map_list:
                    if key in x_str:
                        return target_name
                unmatched_names.append(x_str)
                return "unmatched"
            # 执行映射
            mapped_series = data_series.apply(map_func)
            # 检查是否有未匹配项
            if unmatched_names:
                unique_unmatched = list(set(unmatched_names))
                error_msg = f"{type}名称映射失败，未匹配的名称：{','.join(unique_unmatched)}\n请更新映射表后重试！"
                self._log(error_msg, 'error')
                raise NameMappingError(error_msg)
            return mapped_series
        
        try:
            self._log("正在映射单元名称...")
            df_above['unit'] = map_fuzzy_names(df_above['unit'], unit_map_list, '百万以上单元')
            df_below['unit'] = map_fuzzy_names(df_below['unit'], unit_map_list, '百万以下单元')
            df_contract['unit'] = map_fuzzy_names(df_contract['unit'], unit_map_list, '签约情况单元')
            self._log("成功映射单元名称！")

            self._log("正在映射行业名称...")
            df_above['industry'] = map_fuzzy_names(df_above['industry'], industry_map_list, '百万以上行业')
            df_below['industry'] = map_fuzzy_names(df_below['industry'], industry_map_list, '百万以下行业')
            df_contract['industry'] = map_fuzzy_names(df_contract['industry'], industry_map_list, '签约情况行业')
            self._log("成功映射行业名称！")
        except NameMappingError as e:
            # 映射失败，终止整个流程
            self._update_status("映射名称失败")
            messagebox.showerror("映射失败", str(e))
            return

        # 5. 检查签约情况表中商机编码
        # 5.1 是否在全量商机编码中
        all_bizcode_array = df_bizcode['biz_code'].dropna().unique() # unique函数返回numpy.ndarray
        df_contract['exists'] = df_contract['biz_code'].isin(all_bizcode_array)
        # 5.2 查询商机编码对应的预计签约时间列'date'
        df_contract_merged = df_contract.merge(
            df_bizcode,
            on='biz_code',
            how='left'
        )
        # 5.3 校验实际签约数据是否是宽口径下合规的
        # 第一种：商机编码存在，且计划签约日期为空
        # 第二种：商机编码存在，预计签约时间在实际签约日期的正负三个月内
        df_contract_merged['strict'] = (
            df_contract_merged['exists'] & (
                df_contract_merged['date'].isna() |
                (
                    (df_contract_merged['date'] >= df_contract_merged['sign_date'] - pd.DateOffset(months=3)) &
                    (df_contract_merged['date'] <= df_contract_merged['sign_date'] + pd.DateOffset(months=3))
                )
            )
        )
        self._log("成功标记窄口径签约数据！")

        # 6. 保存预处理之后的数据到文件，留待后续参考
        def save_prehandled_data(file_list):
            prehandled_output_dir = os.path.join(self.base_dir, "预处理后的数据")
            now_str = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
            current_prehandles_dir = os.path.join(prehandled_output_dir, now_str)
            os.makedirs(current_prehandles_dir, exist_ok=True)

            for origin_filepath, sheet_name, df in file_list:
                file_name = os.path.basename(origin_filepath)
                new_filepath = re.sub(
                    r'(\.xlsx|\.xls)$', 
                    fr'_预处理_{now_str}\1',
                    file_name,
                    flags=re.IGNORECASE
                    )
               
                abs_new_filepath = os.path.join(current_prehandles_dir, new_filepath)
                try:
                    df.to_excel(
                        abs_new_filepath,
                        sheet_name = sheet_name,
                        index = False
                    )
                    self._log(f"预处理后的数据文件已保存至: \n{abs_new_filepath}")
                except Exception as e:
                    self._log(f"保存预处理数据失败：\n{traceback.format_exc()}", level="error")
                    self._update_status("保存预处理数据失败")
                    messagebox.showerror("失败", f"保存预处理数据失败：\n{str(e)}")
                    return False
            return True
        
        prehandled_data_file_list = [
            (self.biz_above_million_filepath.get(), self.above_million_sheet, df_above),
            (self.biz_below_million_filepath.get(), self.below_million_sheet, df_below),
            (self.contract_status_filepath.get(), self.contract_status_sheet, df_contract_merged)
        ]  
        if not save_prehandled_data(prehandled_data_file_list):
            return

        # ==================================== 数据汇总 ====================================
        self._update_status("数据汇总")
        self._log("正在进行数据汇总...")
        # 1. 合并
        df_all_biz = pd.concat([df_above, df_below], ignore_index=True, sort=False) # ignore_index重置索引，sort=False避免列名排序
        # 2. 函数
        def unified_count_sum_fun(df, groupby_col, prefix):
            """
            统一的计数和求和函数
            """
            return df.groupby(groupby_col).agg(
                **{f"{prefix}量": (groupby_col, "count")},
                **{f"{prefix}金额": ("money", "sum")}
                )
        
        def unified_stat_by_period_money(df_all_biz, df_contract_filterd, groupby_col, by_month: bool = False):
            """
            可以指定按单元/行业汇总，并分金额范围统计
            @param df_all_biz: 商机清单，包含百万以上、百万以下
            @param df_contract_filterd: 实际签约情况，按不同口径筛选过
            @param groupby_col: 指定按单元/行业汇总
            @param by_month: 区分是按指定时间段统计，还是按月，
            """
            # 如果是按月统计，转换为具体时间
            if by_month:
                start_date = pd.to_datetime(f"{self.stat_year.get()}-{self.stat_month.get()}-01")
                end_date = start_date + MonthEnd(1)
            else:
                start_date = self.start_date.get()
                end_date = self.end_date.get()
            
            df_plan_period = df_all_biz[
                df_all_biz['date'].between(start_date, end_date)
            ]
            
            # 由于百万以上商机预计签约日期只到月，按时间段统计时需要参考“是否目标时间段”这一列
            # 百万以上和百万以下，分开统计
            df_below_period = df_plan_period[~df_plan_period['above_flag']].copy() # between默认是闭区间，所以会包含结束日期，符合需求
            df_below_period_less50 = df_below_period[df_below_period['money'] < 50]
            df_below_period_50to100 = df_below_period[df_below_period['money'].between(50, 100, inclusive='left')]
            df_agg_below_period_less50 = unified_count_sum_fun(df_below_period_less50, groupby_col, '五十万以下计划签约')
            df_agg_below_period_50to100 = unified_count_sum_fun(df_below_period_50to100, groupby_col, '五十至百万计划签约')
            if by_month:
                df_above_period = df_plan_period[df_plan_period['above_flag']]
            else:
                df_above_period = df_all_biz[(df_all_biz['above_flag']) & (df_all_biz['target'] == "是")].copy()
            df_agg_above_period = unified_count_sum_fun(df_above_period, groupby_col, '百万以上计划签约')
            # 计划签约情况
            df_agg_below_period = unified_count_sum_fun(df_below_period, groupby_col, '百万以下计划签约') # 百万以下
            # # merge百万以下、百万以上的整体计划签约情况
            df_agg_plan_period = df_agg_below_period.join(
                df_agg_above_period,
                how='outer'
            ).fillna(0)
            # # 总计 = 百万以上+百万以下
            df_agg_plan_period['计划签约量'] = df_agg_plan_period['百万以上计划签约量'] + df_agg_plan_period['百万以下计划签约量']
            df_agg_plan_period['计划签约金额'] = df_agg_plan_period['百万以上计划签约金额'] + df_agg_plan_period['百万以下计划签约金额']
            # # 删除中间列
            df_agg_plan_period = df_agg_plan_period[['计划签约量', '计划签约金额']].copy()
            
            # 实际签约情况
            df_contract_period = df_contract_filterd[df_contract_filterd['sign_date'].between(start_date, end_date)].copy()
            df_agg_contract_period = unified_count_sum_fun(df_contract_period, groupby_col, '实际签约')
            df_contract_period_less50 = df_contract_period[df_contract_period['money'] < 50]
            df_contract_period_50to100 = df_contract_period[df_contract_period['money'].between(50, 100, inclusive='left')]
            df_contract_period_more100 = df_contract_period[df_contract_period['money'] >= 100]
            df_agg_contract_period_less50 = unified_count_sum_fun(df_contract_period_less50, groupby_col, '五十万以下实际签约')
            df_agg_contract_period_50to100 = unified_count_sum_fun(df_contract_period_50to100, groupby_col, '五十至百万实际签约')
            df_agg_contract_period_more100 = unified_count_sum_fun(df_contract_period_more100, groupby_col, "百万以上实际签约")

            # 横向合并所有统计表（列拼接，外连接保证所有单元都保留）
            df_join = df_agg_plan_period.join(
                [
                    df_agg_contract_period,
                    df_agg_below_period_less50,
                    df_agg_contract_period_less50,
                    df_agg_below_period_50to100,
                    df_agg_contract_period_50to100,
                    df_agg_above_period,
                    df_agg_contract_period_more100
                    ],
                how='outer'
                ).fillna(0)
            # 按指定顺序进行行排序，不存在的单元自动补充0
            row_index = self.unit_order if groupby_col == 'unit' else self.industry_order if groupby_col == 'industry' else None
            df_stat = df_join.reindex(index=row_index, fill_value=0)

            # 转签率 = 实际签约量 / 计划签约量
            # df.apply() 默认给函数依次传入每列数据，数据类型是Series，带索引
            # df.apply(axis=1) 依次传入整行
            # df['列'].apply() 传入单列的每个元素
            # df.applymap() 作用在全表每个单元格
            # df_stat['转签率'] =  df_stat.apply(
            #     lambda x: "{:.2%}".format(x['实际签约量'] / x['计划签约量'])
            #     if x['计划签约量'] != 0 else 0,
            #     axis=1
            # )
            df_stat['转签率'] = 0 # 先赋值为0，在写入excel时，重新计算
            # 按指定顺序重新排布列索引
            df_stat = df_stat.reindex(columns=self.col_order)
            # 给所有列名加前缀
            prefix = "month_" if by_month else "peroid_"
            df_stat = df_stat.add_prefix(prefix)
            return df_stat

        def stat_standard_pipline(df_all_biz, df_contract_merged, groupby_col, strict_mode: bool):
            """
            统一的统计函数，可指定按照宽口径/窄口径、单元/行业维度进行统计
            """
            # 窄口径筛选
            if strict_mode:
                df_contract_filterd = df_contract_merged[df_contract_merged['strict']].copy()
            else:
                df_contract_filterd = df_contract_merged.copy() # 宽口径不筛选，直接使用原表
            mode_logtext = "窄口径" if strict_mode else "宽口径"
            groupby_logtext = "单元" if groupby_col=="unit" else "行业" if groupby_col=="industry" else None
            self._log(f"正在按照{mode_logtext}-{groupby_logtext}维度进行统计...")
            # 整体计划签约情况统计
            df_stat_overall_plan = unified_count_sum_fun(df_all_biz, groupby_col, "整体计划签约")
            df_stat = df_stat_overall_plan.join([unified_count_sum_fun(df_all_biz[df_all_biz['above_flag']], groupby_col, "整体百万以上计划签约")], how='outer').fillna(0)
            # 按指定顺序进行行排序，不存在的单元自动补充0
            row_index = self.unit_order if groupby_col == 'unit' else self.industry_order if groupby_col == 'industry' else None
            df_stat = df_stat.reindex(index=row_index, fill_value=0).fillna(0)
            self._log("整体计划签约情况统计完成！")
            # 指定时间段统计
            df_stat_by_period = unified_stat_by_period_money(df_all_biz, df_contract_filterd, groupby_col)
            self._log("指定时间段统计完成！")
            # 指定月统计
            df_stat_cur_month = unified_stat_by_period_money(df_all_biz, df_contract_filterd, groupby_col, by_month=True)
            self._log("指定月统计完成！")
            df_stat = df_stat.join([df_stat_by_period, df_stat_cur_month], how='outer').fillna(0)
            # 统计随后三个月
            current_month_fisrt_day  = pd.to_datetime(f"{self.stat_year.get()}-{self.stat_month.get()}-01")
            for idx in range(1, 4):
                start_dt = current_month_fisrt_day + pd.DateOffset(months=idx)
                end_dt = start_dt + MonthEnd(1)
                df_all_biz_filtered = df_all_biz[df_all_biz['date'].between(start_dt, end_dt)].copy()
                df_stat = df_stat.join([unified_count_sum_fun(df_all_biz_filtered, groupby_col, f"{start_dt.month}月计划签约")], how='outer').fillna(0)
            self._log("随后三月统计完成！")

             # 计算[合计]行
            # to_frame将Series强行包装成二维DataFrame
            # pandas.core.series.Series
            # A    5
            # B    7
            # C    9
            # pandas.core.frame.DataFrame
            #    0
            # A  5
            # B  7
            # C  9
            df_total = df_stat.sum().to_frame().T
            df_total.index = [f"{'单位' if groupby_col == 'unit' else '行业' if groupby_col == 'industry' else ''}合计"]
            # 合并合计行 + 明细行
            df_final = pd.concat([df_total, df_stat], ignore_index=True, sort=False)

            return df_final
        
        def filter_top_10(df_all_biz):
            """
            筛选指定月计划签约金额的top10商机
            """
            self._log("正在筛选指定月商机top10...")
            start_dt = pd.to_datetime(f"{self.stat_year.get()}-{self.stat_month.get()}-01")
            end_dt = start_dt + MonthEnd(1)
            df_filt = df_all_biz[df_all_biz['date'].between(start_dt, end_dt)].copy()
            df_top_10 = df_filt.sort_values(by='money', ascending=False).head(10).copy()
            df_top_10 = df_top_10.reindex(columns=['unit', 'industry', 'name', 'money', 'date'])
            df_top_10['date'] = df_top_10['date'].dt.strftime("%Y%m")
            self._log("指定月商机top10筛选完成！")
            return df_top_10

        def save_to_report_file(df_unit_stat, df_industry_stat, df_top_10, strict_mode: bool):
            """
            将统计结果进报告模板文件里并保存
            @report_filepath 最终要填写的报告模板文件地址
            @param df_unit 按单元维度统计的商机计划与签约信息
            @param df_industry 按行业维度统计的商机计划与签约信息
            @param df_top_10 指定月top10商机
            """
            def write_df_to_template_sheet(ws, df, start_row, start_col):
                for row_idx, row_data in enumerate(df.itertuples(index=False, name=None)): # name=None不生产命名元组，直接返回普通元组
                    for col_idx, cell_value in enumerate(row_data):
                        target_row = start_row + row_idx
                        target_col = start_col + col_idx
                        target_cell = ws.cell(target_row, target_col)
                        target_cell.value = cell_value
                        # 处理转签率这一列，转签率总和列单独算，计划签约为0时，转签率置为"/"
                        if col_idx == 6 or col_idx == 23:
                            if ws.cell(target_row, target_col - 2).value == 0:
                                target_cell.value = "/"
                            else:
                                target_cell.value = ws.cell(target_row, target_col - 1).value / ws.cell(target_row, target_col - 2).value

            self._log("开始加载通报模板...")
            self._update_status("生成报告")
            try:
                wb = openpyxl.load_workbook(
                    filename=self._report_template_filepath,
                    data_only=False # 保留公式、格式
                    )
            except Exception as e:
                self._log(f"加载通报模板失败：\n{traceback.format_exc()}", "error")
                self._update_status("报告生成失败")
                messagebox.showerror("路径无效", f"无法加载通报模板！")
                return False
            
            target_sheet_name = "1.有效商机转签情况"
            if target_sheet_name not in wb.sheetnames:
                self._log(f"模板中未找到目标Sheet: {target_sheet_name}", "error")
                self._update_status("报告生成失败")
                messagebox.showerror("Sheet不存在", f"模板中未找到目标Sheet: {target_sheet_name}")
                return False
            
            ws = wb[target_sheet_name]
            self._log(f"成功加载目标Sheet: {target_sheet_name}!")

            # 更新表头
            period_st = pd.to_datetime(self.start_date.get())
            period_ed = pd.to_datetime(self.end_date.get())
            period_col_label = f'{period_st.year}年{period_st.month}月{period_st.day}日-{period_ed.year}年{period_ed.month}月{period_ed.day}日'
            ws.cell(2, 6).value = period_col_label
            ws.cell(41, 6).value = period_col_label

            cur_month = pd.to_datetime(f"{self.stat_year.get()}-{self.stat_month.get()}")
            cur_month_col_label = f'{cur_month.year}年{cur_month.month}月'
            ws.cell(2, 23).value = cur_month_col_label
            ws.cell(41, 23).value = cur_month_col_label

            col_offset = 0
            for _ in range(0, 3):
                cur_month += pd.DateOffset(months=1)
                col_label = f'{cur_month.year}年{cur_month.month}月'
                ws.cell(2, 40 + col_offset).value = col_label
                ws.cell(41, 40 + col_offset).value = col_label
                col_offset += 2

            # 定义三部分的起始位置，从1开始计数
            write_config = {
                # 单元维度
                "unit": {
                    "start_row": 4,
                    "start_col": 2
                },
                # 行业维度
                "industry": {
                    "start_row": 43,
                    "start_col": 2
                },
                # top10
                "top_10": {
                    "start_row": 60,
                    "start_col": 1
                }
            }

            # 写入指定位置
            if not df_unit_stat.empty:
                self._log("开始写入单元维度统计数据...")
                write_df_to_template_sheet(
                    ws=ws,
                    df=df_unit_stat,
                    start_row=write_config['unit']['start_row'],
                    start_col=write_config['unit']['start_col']
                )
            else:
                self._log("单元维度统计数据为空", "warning")
            if not df_industry_stat.empty:
                self._log("开始写入行业维度统计数据...")
                write_df_to_template_sheet(
                    ws=ws,
                    df=df_industry_stat,
                    start_row=write_config['industry']['start_row'],
                    start_col=write_config['industry']['start_col']
                )
            else:
                self._log("行业维度统计数据为空", "warning")
            if not df_top_10.empty:
                self._log("开始写入商机top10数据...")
                write_df_to_template_sheet(
                    ws=ws,
                    df=df_top_10,
                    start_row=write_config['top_10']['start_row'],
                    start_col=write_config['top_10']['start_col']
                )
            else:
                self._log("商机top10数据为空", "warning")
            # 保存文件
            filename = f"商机报告_{'窄口径' if strict_mode else '宽口径'}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.base_dir, filename)
            try:
                wb.save(filepath)
                wb.close()
                self._log(f"报告已生成！文件保存至：\n{filepath}")
                self._update_status("报告生成成功")
                messagebox.showinfo("生成成功", f"报告已生成！文件保存至：\n{filepath}")
                return True
            except Exception as e:
                wb.close()
                self._log(f"报告文件保存失败: \n{traceback.format_exc()}", "error")
                self._update_status("报告生成失败")
                messagebox.showerror("生成失败", f"报告生成失败：\n{str(e)}")
                return False
            
        # 3. 调用函数进行统计, 并保存结果到文件
        df_top_10 = filter_top_10(df_all_biz)
        # 宽口径
        if self.loose_mode.get():
            df_loose_unit_stat = stat_standard_pipline(df_all_biz, df_contract_merged, groupby_col='unit', strict_mode=False)
            df_loss_industry_stat = stat_standard_pipline(df_all_biz, df_contract_merged, groupby_col='industry', strict_mode=False)
            if not save_to_report_file(df_loose_unit_stat, df_loss_industry_stat, df_top_10, strict_mode=False):
                return
        if self.strict_mode.get():
            df_strict_unit_stat = stat_standard_pipline(df_all_biz, df_contract_merged, groupby_col='unit', strict_mode=True)
            df_strict_industry_stat = stat_standard_pipline(df_all_biz, df_contract_merged, groupby_col='industry', strict_mode=True)
            save_to_report_file(df_strict_unit_stat, df_strict_industry_stat, df_top_10, strict_mode=True)

    def run_biz_analysis_workflow(self):
        try:
            # 1. 读取文件并检查Sheet名、列名
            df_set = self._open_source_table()
            if not df_set:
                return
            # 2. 统计
            self._summarize_by_unit_and_industry(*df_set)
        except Exception as e:
            self._log(f"执行统计流程出错：\n{traceback.format_exc()}", 'error')
            self._update_status("统计流程出错")
            messagebox.showerror("错误", f"执行统计流程出错：\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk() # Tk()是创建一个Tkinter应用程序的主窗口对象，所有的组件都要放到这个主窗口上，
    app = BizReportApp(root)
    # mainloop()是Tkinter应用程序的事件循环，负责监听用户的操作并做出相应，必须在创建完所有组建后调用
    root.mainloop()