from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def make_directory_catalog(folder: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = '目录'
    ws.merge_cells('A1:F1')
    ws['A1'] = '目录清单'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['序号', '名称', '文件格式', '文件大小', '是否经省公司分管领导审核', '备注']
    ws.append(headers)

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alignment = Alignment(horizontal='center', vertical='center')
    font_text = Font(name='黑体', size=11)
    font_title_header = Font(name="黑体", size=14, bold=True)

    # 定义计算文件夹大小的函数，递归计算文件夹内所有文件的大小之和，单位是字节
    def get_folder_size(folder_path):
        total_size = 0
        for item in folder_path.iterdir():
            if item.is_file():
                total_size += item.stat().st_size
            elif item.is_dir():
                total_size += get_folder_size(item)
        return total_size
    
    # # 提取序号作为排序依据，无序号放最后
    # def sort_key(file: Path):
    #     first_dot_pos = str(file.name).find('.')
    #     if first_dot_pos != -1 and str(file.name)[:first_dot_pos].isdigit():
    #         return int(file.name[:first_dot_pos])
    #     return float('inf')  # 将无序号的文件放在最后

    # data rows
    data_rows_list = []
    for file in sorted(folder.iterdir()): # sorted按照文件名排序，iterdir()返回目录下的所有文件和文件夹的Path对象, iterdir()默认只返回当前目录下的内容，不会递归进入子目录。
        # 跳过所有已存在的"0.目录"文件
        if file.is_file() and file.name == "0.目录.xlsx": # file.name包含文件名和扩展名，file.stem只包含文件名不包含扩展名
            continue
        first_dot_pos = file.name.find('.')
        possible_index_str = file.name[: first_dot_pos] if first_dot_pos != -1 else '' 
        if possible_index_str and possible_index_str.isdigit(): # isdigit()方法检查字符串是否只包含数字字符，如果是数字则返回True，否则返回False
            index = possible_index_str
            name_remv_order = file.name[first_dot_pos+1:] # 从第一个点之后开始截取文件名，直到末尾
        else:
            index = ''
            name_remv_order = file.name
        if file.is_file():
            # 最后一个点拆分名称和后缀
            last_dot_pos = name_remv_order.rfind('.')
            if last_dot_pos != -1:
                file_name_without_suffix = name_remv_order[:last_dot_pos] 
                suffix = name_remv_order[last_dot_pos+1:].lower() # 从最后一个点之后开始截取后缀，直到末尾
            else:
                file_name_without_suffix = name_remv_order
                suffix = ''
            file_size = file.stat().st_size # stat()函数返回一个包含文件状态信息的对象，st_size属性表示文件的大小（以字节为单位）
            # 根据文件大小自动显示单位
            if file_size < 1024:
                size_str = f"{file_size}B"
            elif file_size < 1024 * 1024:
                size_str = f"{round(file_size / 1024, 2)} KB" # 2表示保留两位小数
            else:
                size_str = f"{round(file_size / (1024 * 1024), 2)} MB" # 2表示保留两位小数
            data_rows_list.append([index, file_name_without_suffix, suffix, size_str, '否', ''])
            
        if file.is_dir():
            # 计算文件夹的大小
            folder_size = get_folder_size(file) # 计算文件夹大小，单位是字节
            # 根据文件夹大小自动显示单位
            if folder_size < 1024:
                size_str = f"{folder_size}B"
            elif folder_size < 1024 * 1024:
                size_str = f"{round(folder_size / 1024, 2)} KB" # 2表示保留两位小数
            else:
                size_str = f"{round(folder_size / (1024 * 1024), 2)} MB" # 2表示保留两位小数

            data_rows_list.append([index, name_remv_order, '文件夹', size_str, '否', ''])

    # 按照序号列进行排序，无序号的放在最后
    data_rows_list.sort(key=lambda x: int(x[0]) if x[0].isdigit() else float('inf')) # x[0]是序号列，如果是数字则转换为整数进行排序，如果不是数字则返回正无穷大，使其排在最后
    # 将数据行写入工作表
    for data_row in data_rows_list:
        ws.append(data_row)
    # style header cells
    max_row = ws.max_row
    max_col = ws.max_column
    # title_cell = ws.cell(1,1)
    # title_cell.font = font_title_header
    # title_cell.border = border
    # for c in range(1, max_col + 1):
    #     header_cell = ws.cell(2, c)
    #     header_cell.font = font_title_header
    #     header_cell.border = border
    #     header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # all content cells
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if r == 1 or r ==2:
                cell.font = font_title_header
            else:
                cell.font = font_text
            cell.border = border
            cell.alignment = alignment
    # 所有列设置固定宽度22 
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 22  # 固定列宽为22, column_dimensions是openpyxl中用于设置列属性的对象，get_column_letter函数将列索引转换为Excel列字母，例如1对应A，2对应B，以此类推。
        # 至少设置第三行的格式
        cell = ws.cell(row=3, column=c)
        cell.font = font_text
        cell.border = border
        cell.alignment = alignment

    # # auto column width
    # for c in range(1, max_col + 1):
    #     col_letter = ws.cell(row=1, column=c).column_letter
    #     max_length = 0
    #     for r in range(1, max_row + 1):
    #         value = ws.cell(row=r, column=c).value
    #         if value is None:
    #             continue
    #         max_length = max(max_length, len(str(value)))
    #     ws.column_dimensions[col_letter].width = max_length + 2

    target = folder / '0.目录.xlsx'
    wb.save(str(target)) # 覆盖保存，如果文件已存在则覆盖，如果不存在则创建新文件
    return target

def main():
    root_H = Path('F:/')
    first_level_folder = root_H / '6.西安-近5年中涉及调用原子能力的合同的资料'
    if not first_level_folder.exists():
        print(f'未找到一级目录: {first_level_folder}')
        return  
    
    # 递归遍历所有子目录
    def traverse_dir(folder: Path):
        # 先处理子文件夹，再处理当前文件夹，保证先生成子目录的目录文件，再生成当前目录的目录文件
        for item in folder.iterdir():
            if item.is_dir():
                traverse_dir(item) # 递归处理子目录
        # 处理当前目录，生成目录文件
        print(f'正在处理目录: {folder}')
        make_directory_catalog(folder)
        
    # # 生成一级目录文件
    # make_directory_catalog(first_level_folder)
    # for second_level_folder in first_level_folder.iterdir(): 
    #     # 跳过新生成的一级目录文件
    #     if second_level_folder.is_dir():
    #         # 生成二级目录文件
    #         print(f'正在处理目录: {second_level_folder}')
    #         make_directory_catalog(second_level_folder)
    #         for third_level_folder in second_level_folder.iterdir():
    #             # 跳过新生成的二级目录文件
    #             if third_level_folder.is_dir():
    #                 # 生成三级目录文件
    #                 print(f'正在处理目录: {third_level_folder}')
    #                 make_directory_catalog(third_level_folder)
    traverse_dir(first_level_folder)
    print('各级目录生成完毕！')

if __name__ == '__main__':
    main()