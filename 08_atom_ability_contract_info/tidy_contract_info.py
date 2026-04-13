import os
import shutil
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


def sanitize_name(name: str) -> str:
    # avoid invalid windows chars
    return re.sub(r'[<>:"/\\|?*]', '_', name)


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    base = path.stem # stem是不带后缀的文件名部分
    suffix = path.suffix # suffix是文件的后缀部分，包括点，例如'.xlsx'
    parent = path.parent # parent是路径的父目录部分
    i = 1
    while True:
        candidate = parent / f"{base}_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def copy_with_unique(src: Path, dst_dir: Path, base_name: str) -> Path:
    # 函数目前未被使用，保留以备后续需要更灵活的命名方式时使用
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst_path = dst_dir / f"{base_name}{src.suffix}"
    dst_path = unique_path(dst_path)
    shutil.copy2(src, dst_path) # shutil是一个高级文件操作模块，copy2函数会复制文件并尽可能保留元数据（如修改时间等）
    return dst_path


def copy_file_with_unique(src: Path, dst_dir: Path) -> Path:
    dst_dir.mkdir(parents=True, exist_ok=True) # parents=True确保父目录存在，exist_ok=True避免已存在时报错
    dst_path = dst_dir / src.name # src.name是文件名（不含路径）
    dst_path = unique_path(dst_path)
    shutil.copy2(src, dst_path) # # shutil是一个高级文件操作模块，copy2函数会复制文件并尽可能保留元数据（如修改时间等）
    return dst_path


def normalize_text(text: str) -> str:
    import unicodedata
    text = unicodedata.normalize('NFKC', text or '') # NFKC是Unicode的一种规范化形式，能将全角字符转换为半角，兼容性分解等，有助于统一文本格式。
    # Filter to keep only letters and numbers (Chinese, Latin, digits)
    filtered = []
    for ch in text:
        cat = unicodedata.category(ch) # category函数返回字符的Unicode类别，例如'L'开头表示字母，'N'开头表示数字，'Z'开头表示分隔符等。
        if cat.startswith('L') or cat.startswith('N'):
            filtered.append(ch)
    return ''.join(filtered).lower()


def find_first_matching_folder(root: Path, contract_name: str):
    contract_name_norm = normalize_text(contract_name or '')
    if not contract_name_norm:
        return None
    for dirpath, dirnames, _ in os.walk(root):
        for dirname in dirnames:
            dirname_norm = normalize_text(dirname)
            if contract_name_norm in dirname_norm:
                return Path(dirpath) / dirname
    return None


def find_files_by_keywords(root: Path, keywords):
    matches = []
    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:
            name_lower = filename.lower()
            for kw in keywords:
                if kw.lower() in name_lower:
                    matches.append(Path(dirpath) / filename)
                    break 
    return matches


def find_first_file_by_keywords(root: Path, keywords):
    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:
            name_lower = filename.lower()
            for kw in keywords:
                if kw.lower() in name_lower:
                    return Path(dirpath) / filename
    return None


def find_first_file_by_any(root: Path, keywords):
    for dirpath, _, filenames in os.walk(root):
        for filename in filenames:
            filename_norm = normalize_text(filename)
            for kw in keywords:
                kw_norm = normalize_text(kw)
                if kw_norm and kw_norm in filename_norm:
                    return Path(dirpath) / filename
    return None


def make_directory_catalog(folder: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = '目录'
    ws.merge_cells('A1:F1')
    ws['A1'] = '目录清单'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ['序号', '名称', '文件格式（zip、rar、word、pdf、ppt等）', '文件大小', '是否经分管领导审核', '备注']
    ws.append(headers)

    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alignment = Alignment(horizontal='center', vertical='center')
    font_text = Font(name='黑体', size=14)
    font_title = Font(name="黑体", size=20, bold=True)

    # data rows
    index = 1
    for file in sorted(folder.iterdir()): # sorted按照文件名排序，iterdir()返回目录下的所有文件和文件夹的Path对象, iterdir()默认只返回当前目录下的内容，不会递归进入子目录。
        if file.is_file():
            name = file.name
            suffix = file.suffix.lstrip('.').lower() or '' # lstrip全称是去除左边的字符,lower()转换为小写，如果没有后缀则返回空字符串
            size = file.stat().st_size # stat()函数返回一个包含文件状态信息的对象，st_size属性表示文件的大小（以字节为单位）
            # 根据文件大小自动显示单位
            if size < 1024:
                size_str = f"{size}B"
            elif size < 1024 * 1024:
                size_str = f"{round(size / 1024)}kb"
            else:
                size_str = f"{round(size / (1024 * 1024))}mb"
            ws.append([index, name, suffix, size_str, '', ''])
            index += 1
        if file.is_dir():
            name = file.name # name返回文件夹的名称
            ws.append([index, name, '文件夹', '', '', ''])
            index += 1

    # style header cells
    max_row = ws.max_row
    max_col = ws.max_column
    title_cell = ws.cell(1,1)
    title_cell.font = font_title
    title_cell.border = border
    for c in range(1, max_col + 1):
        header_cell = ws.cell(2, c)
        header_cell.font = font_title
        header_cell.border = border
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # all content cells
    for r in range(3, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_text
            cell.border = border
            cell.alignment = alignment
    # 所有列设置固定宽度22 
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 22  # 固定列宽为22, column_dimensions是openpyxl中用于设置列属性的对象，get_column_letter函数将列索引转换为Excel列字母，例如1对应A，2对应B，以此类推。
        # 至少设置第三行的格式
        cell = ws.cell(row=3, column=c)
        cell.font = font_text
        cell.border = border
        cell.alignment = alignment

    # # auto column width
    # for c in range(1, max_col + 1):
    #     col_letter = ws.cell(row=1, column=c).column_letter
    #     max_length = 0
    #     for r in range(1, max_row + 1):
    #         value = ws.cell(row=r, column=c).value
    #         if value is None:
    #             continue
    #         max_length = max(max_length, len(str(value)))
    #     ws.column_dimensions[col_letter].width = max_length + 2

    target = folder / '0.目录.xlsx'
    wb.save(str(target))
    return target

def main():
    root_H = Path('H:/')
    source_xlsx = root_H / '合同_大于十万.xlsx'
    if not source_xlsx.exists():
        print(f'未找到合同源文件: {source_xlsx}')
        return

    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    target_xlsx = root_H / f'{source_xlsx.stem}_{now}.xlsx'
    shutil.copy2(str(source_xlsx), str(target_xlsx))

    df = pd.read_excel(str(target_xlsx), dtype=str)
    df = df.astype(object)

    for col in ['1.合同签报', '2.订单截图', '3.调用依据', '4.验收材料']:
        if col not in df.columns:
            df[col] = '/'

    atomic_root = root_H / '原子能力'
    acceptance_root = root_H / '全部验收报告'
    output_root = root_H / '西安-近5年中涉及调用原子能力的合同的资料'
    output_root.mkdir(parents=True, exist_ok=False)

    processed_folders = []

    for idx, row in df.iterrows():
        serial = str(row.get('序号', '')).strip() or str(idx + 1)
        contract_code = str(row.get('合同编码', '')).strip()
        contract_name = str(row.get('合同名称', '')).strip()
        if not contract_code and not contract_name:
            continue

        second_folder_name = sanitize_name(f"{serial}.{contract_code}+{contract_name}")
        second_folder = output_root / second_folder_name
        second_folder.mkdir(parents=True, exist_ok=True)

        processed_folders.append(second_folder)

        # 1. in 原子能力资料文件夹 find folder by contract_name
        found_folder = None
        if atomic_root.exists() and contract_name:
            found_folder = find_first_matching_folder(atomic_root, contract_name)

        # define categories
        category_rules = [
            ('1.合同签报', ['合同', '协议']),
            ('2.订单截图', ['订单截图', '调用截图', '下单截图']),
            ('3.调用依据', ['纪要', '申请单', '需求表', '调用单', '申请表', '需求单', '调用表', '情况说明']),
            ('4.验收材料', []),
        ]

        # pre-create category folders
        category_dirs = {}
        for cat, _ in category_rules:
            third_folder = second_folder / cat
            third_folder.mkdir(parents=True, exist_ok=True)
            category_dirs[cat] = third_folder

        # search and copy for first three categories from 原子能力资料文件夹
        if found_folder:
            for cat, keywords in category_rules[:3]:
                matched = []
                for dirpath, _, filenames in os.walk(found_folder):
                    for filename in filenames:
                        low = filename.lower()
                        if any(kw.lower() in low for kw in keywords):
                            matched.append(Path(dirpath) / filename)
                if matched:
                    df.at[idx, cat] = '是' # at的用法：df.at[row_index, column_name] = value，直接定位到DataFrame的单元格进行赋值，效率较高。
                    for fpath in matched:
                        copy_file_with_unique(fpath, category_dirs[cat])

        # 2. 验收材料 in 全部验收报告文件夹
        if acceptance_root.exists() and (contract_code or contract_name):
            found_accept = None
            if contract_code:
                found_accept = find_first_file_by_any(acceptance_root, [contract_code])
            if not found_accept and contract_name:
                found_accept = find_first_file_by_any(acceptance_root, [contract_name])

            if found_accept:
                df.at[idx, '4.验收材料'] = '是'
                copy_file_with_unique(found_accept, category_dirs['4.验收材料'])

        # 生成每个三级分类文件夹下的 0.目录.xlsx
        # for dir_path in category_dirs.values(): 
        #     make_directory_catalog(dir_path)

        # 生成每个二级分类文件夹下的 0.目录.xlsx
        # make_directory_catalog(second_folder) 

    # 生成根目录下的 0.目录.xlsx（目录脚本）
    # make_directory_catalog(output_root)

    # save contract excel
    df.to_excel(str(target_xlsx), index=False)
    print(f'处理完成，输出文件: {target_xlsx}')


if __name__ == '__main__':
    main()
