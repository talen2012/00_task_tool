import time
import random
import openpyxl
from copy import copy
from openpyxl.styles import Font, colors
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.common.exceptions import NoSuchElementException, TimeoutException, ElementNotInteractableException

# @Time    : 2025/09/01 11:32
# @Author  : talen
# @File    : company_info_collector.py

def manual_login(driver, wait_time=3):
    # 检查是否有用户问候语
    try:
        WebDriverWait(driver, wait_time).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "username-company"))
        )
        print("检测到已登录状态，跳过登录环节。")
    except TimeoutException:
        while True:
            user_input = input("未登录或首页未加载完整，处理后输入ok继续...\n")
            if user_input.strip().lower() == "ok":
                break
def open_and_manual_login(driver, target_url):
    """
    打开网站并手动登录
    """
    driver.get(target_url)
    print(f"正在打开目标网站：{target_url}...")
    time.sleep(1)
    manual_login(driver)

def find_exact_company_in_providers(driver, company_name):
    """
    在能力提供方页面查找完全匹配的公司名称，找到后立即返回，未找到则返回None
    :param company_name: 需精确匹配的公司名称
    :param driver: 浏览器驱动实例
    :return: 匹配的公司元素或None
    """
    page_num = 1 # 当前页码
    while True:
        print(f"在第 {page_num} 页查找公司：{company_name}...")

        # 1. 获取当前页的所有公司名称元素（class为goods-item）
        try:
            current_page_companies = WebDriverWait(driver,10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "company-title"))
            )
            time.sleep(random.uniform(1, 2)) # 随机等待1-2秒，模拟人类操作
        except TimeoutException:
            print(f"第 {page_num} 页未加载出公司列表，停止查找")
            return None
        # 2. 遍历当前页公司，检查是否有完全匹配的名称
        for company in current_page_companies:
            # company的text属性只获取了没有被<font>标签包围的部分
            # get_attribute("innerText"或"textContent")可以获得完整文本
            if company.get_attribute("innerText").strip() == company_name : # 去除首尾空格，完全匹配
                print(f"在第 {page_num} 页找到完全匹配的公司：{company_name}")
                return company
            
        # 3. 当前页无匹配，检查是否有下一页可点击
        try:
            # 定位下一页按钮（可点击状态：包含ivu-page-next且不包含disabled）
            next_page_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((
                    By.XPATH, "//li[contains(@class, 'ivu-page-next') and not(contains(@class, 'ivu-page-disabled'))]"
                    ))
            )
            time.sleep(random.uniform(0.5, 1))  # 随机等待0.5-1秒，模拟人类操作
            next_page_btn.click()  # 点击下一页
            page_num += 1
        except (TimeoutException, ElementNotInteractableException):
            print(f"已遍历所有页面，共{page_num} 页，未找到匹配的公司：{company_name}")
            return None
        

def search_target_company(driver, company_name):
    """
    搜索并进入公司能力清单页面
    """
    try:
        # 1. 点击【搜索图标】，触发搜索输入框显示
        # WebDriverWait等待元素可点击，避免因页面加载慢导致找不到元素
        # 10秒超时,until条件成立后继续执行
        # <i data-v-db38ac84="" class="search_Icon ivu-icon ivu-icon-ios-search"></i>
        print(f"点击搜索图标,输入“{company_name}”并回车，等待搜索结果...")
        # 首页
        driver.switch_to.window(driver.window_handles[0])
        search_icon = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "search_Icon"))  # 搜索图标的class（从HTML中提取）
        )
        search_icon.click()
        time.sleep(random.uniform(0.5, 1))

        # 2. 定位【搜索输入框】并输入公司名称
        # XPath说明：找包含class="header_search"的div下的input标签
        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'search_input_box')]//input "))
        )
        search_input.clear()  # 清空输入框（若有默认提示文字）
        search_input.send_keys(company_name)  # 输入公司名称
        time.sleep(random.uniform(1, 2))
        search_input.send_keys(Keys.ENTER)  # 按回车搜索

        # 3. 点击【能力提供方】导航标签
        # 定位逻辑：找包含"能力提供方"文本的div（class包含ivu-tabs-tab...）
        # 同时等待能力提供方标签可点击和goods-content-title元素全部加载
        WebDriverWait(driver, 15).until(
            lambda d: EC.element_to_be_clickable((
                By.XPATH, "//div[contains(@class, 'ivu-tabs-tab') and contains(text(), '能力提供方')]"
            ))(d) and len(WebDriverWait(d, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "goods-item")))) > 0
        )
        ability_provider_tab = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'ivu-tabs-tab') and contains(text(), '能力提供方')]")))
        time.sleep(random.uniform(0.5, 1)) # 随机等待0.5-1秒，模拟人类操作
        ability_provider_tab.click()

        # 4. 逐页查找完全匹配的公司（找到即停）
        target_company = find_exact_company_in_providers(driver, company_name)
        return target_company

    except Exception as e:
        print(f"查找公司能力清单时发生错误：{str(e)}")
        return None

def extract_ability_details(driver, ability_details):
    """
    在能力详情页提取所需信息
    """
    try:
        try:
            # 售前电话和技术支持电话 或者客服电话，需要点击图标
            contact_icon = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "message-back"))
            )
            time.sleep(random.uniform(0.5, 1))
            contact_icon.click()

            # 点击之后，如果登录状态失效，会自动弹出登录界面
            try:
                WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'login_box-button')]"))
                )
                while True:
                    input_str = input("登录状态失效，请重新登录后输入ok继续...\n")
                    if input_str.strip().lower() == "ok":
                        break
                # 再次点击联系方式图标
                # driver.switch_to.window(driver.window_handles[2])
                contact_icon = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, "message-back"))
                )
                time.sleep(random.uniform(0.5, 1))
                contact_icon.click()
            except TimeoutException:
                print("登录状态维持中...")
            try:
                ability_details["售前电话"] = WebDriverWait(driver, 1).until(
                    EC.presence_of_element_located((By.XPATH,
                                                "//div[@class='message-list modal-class']/div[text()='售前电话']/following-sibling::div[@class='text']//div[@class='modal-text']"
                                                ))
                ).text.strip()
            except TimeoutException:
                ability_details["售前电话"] = "/"
            try:
                ability_details["技术支持电话"] = WebDriverWait(driver, 1).until(
                    EC.presence_of_element_located((By.XPATH,
                                                    "//div[@class='message-list modal-class']/div[text()='技术支持电话']/following-sibling::div[@class='text']//div[@class='modal-text']"
                                                    ))
                ).text.strip()
            except TimeoutException:
                ability_details["技术支持电话"] = "/"
            try:
                ability_details["客服电话"] = WebDriverWait(driver, 1).until(
                    EC.presence_of_element_located((By.XPATH, 
                                                    "//div[@class='message-list modal-class']/div[text()='客服电话']/following-sibling::div[@class='text']//span[@class='modal-text']"
                                                    ))
                ).text.strip()
            except TimeoutException:
                ability_details["客服电话"] = "/"

            # 点击评价按钮
            review_btn = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,"//button[contains(@class, 'info-button')]/span[text()='评价']"))
            )
            time.sleep(random.uniform(1, 2))
            review_btn.click()
            # 弹出的对话框选择"满意"
            agree_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "Agree"))
            )
            time.sleep(random.uniform(0.5, 1))
            agree_btn.click()
            # 再点击确定按钮
            confirm_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "(//button[contains(@class, 'ivu-btn-primary')]/span[text()='确定']/parent::button)[2]")) # 先定位到span，再回到父级，匹配到两个，选择第二个
            )
            time.sleep(random.uniform(0.5, 1))
            confirm_btn.click()
        except TimeoutException:
            print("联系方式弹窗异常！")
            ability_details["状态"] = "查询失败"
            return ability_details
        # 能力名称
        # 无需重复提取
        # 能力介绍
        try:
            ability_details["能力介绍"] = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "buyDetailText"))
            ).text.strip()
        except TimeoutException:
            ability_details["能力介绍"] = "/"
        # 能力编码
        try:
            ability_details["能力编码"] = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, 
                                                # following-sibling的意思是选取某个节点之后的兄弟节点
                                                "//span[contains(@class, 'product-specification-label') and text()='能力编码']/following-sibling::span[contains(@class, 'product-specification-span')]"
                                                ))
        ).text.strip()
        except TimeoutException:
            ability_details["能力编码"] = "/"
        # 能力ID
        try:
            ability_details["能力ID"] = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, 
                                                "//span[contains(@class, 'product-specification-label') and text()='能力ID']/following-sibling::span[contains(@class, 'product-specification-span')]"
                                                ))
            ).text.strip()
        except TimeoutException:
            ability_details["能力ID"] = "/"
        # 能力类型
        try:
            ability_details["能力类型"] = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'product-specification-label') and text()='能力类型']/following-sibling::span[contains(@class, 'product-specification-span')]"))
            ).text.strip()
        except TimeoutException:
            ability_details["能力类型"] = "/"
        # 细分类型
        try:
            ability_details["细分类型"] = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'product-specification-label') and text()='细分类型']/following-sibling::span[contains(@class, 'product-specification-span')]"))
            ).text.strip()
        except TimeoutException:
            ability_details["细分类型"] = "/"
        # 能力目录
        try:
            ability_classification = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'product-specification-label') and text()='能力目录']/following-sibling::span[contains(@class, 'product-specification-span')]"))
            ).text.strip()
            ability_details["能力目录一级"] = ability_classification.split("-")[0].strip() if len(ability_classification.split("-")) > 0 else "/"
            ability_details["能力目录二级"] = ability_classification.split("-")[1].strip() if len(ability_classification.split("-")) > 1 else "/"
            ability_details["能力目录三级"] = ability_classification.split("-")[2].strip() if len(ability_classification.split("-")) > 2 else "/"
            ability_details["能力目录四级"] = ability_classification.split("-")[3].strip() if len(ability_classification.split("-")) > 3 else "/"
            ability_details["能力目录五级"] = ability_classification.split("-")[4].strip() if len(ability_classification.split("-")) > 4 else "/"
        except TimeoutException:
            ability_details["能力目录一级"] = "/"
            ability_details["能力目录二级"] = "/"
            ability_details["能力目录三级"] = "/"
            ability_details["能力目录四级"] = "/"
            ability_details["能力目录五级"] = "/"
        # 上架日期
        try:
            ability_details["上架日期"] = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'product-specification-label') and text()='上架日期']/following-sibling::span[contains(@class, 'product-specification-span')]"))
            ).text.strip()
        except TimeoutException:
            ability_details["上架日期"] = "/"
        # 更新日期
        try:
            ability_details["更新日期"] = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'product-specification-label') and text()='更新日期']/following-sibling::span[contains(@class, 'product-specification-span')]"))
            ).text.strip()
        except TimeoutException:
            ability_details["更新日期"] = "/"
        # 能力亮点，最多五个
        try:
            ability_highlights_title = WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "HighlightTitle"))
            )
            ability_highlights_content = WebDriverWait(driver, 5).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "HighlightContent"))
            )
            for i, (title, content) in enumerate(zip(ability_highlights_title, ability_highlights_content)):
                ability_details[f"能力亮点{i+1}"] = f"{title.text.strip()}：{content.text.strip()}"
                if i >= 4 : # 最多五个亮点
                    break
            for i in range(len(ability_highlights_title), 5): # 填充缺失的亮点
                ability_details[f"能力亮点{i+1}"] = "/"
        except TimeoutException: 
            for i in range(5):
                ability_details[f"能力亮点{i+1}"] = "/"

        # 如果除了能力名称，其他字段都没有提取到，说明查询失败
        if all(value == "/" for key, value in ability_details.items() if key != "能力名称" and key != "状态"):
            print(f"提取能力详情失败！")
            ability_details["状态"] = "查询失败"
            return ability_details

        # 流程完整运行
        ability_details["状态"] = "查询成功"
        return ability_details

    except Exception as e:
        print(f"提取能力详情时发生错误")
        ability_details["状态"] = "查询失败"
        return ability_details

def write_single_ability_detail_to_excel(target_file_name, wb, ability_details_list_ws, ability_details):
    """
    将获取到的单个能力详情写入Excel，
    """
    row = [ability_details.get("能力名称", ""),
           ability_details.get("状态", ""),
           ability_details.get("能力介绍", ""),
           ability_details.get("能力编码", ""),
           ability_details.get("能力ID", ""),
           ability_details.get("能力类型", ""),
           ability_details.get("细分类型", ""),
           ability_details.get("能力目录一级", ""),
           ability_details.get("能力目录二级", ""),
           ability_details.get("能力目录三级", ""),
           ability_details.get("能力目录四级", ""),
           ability_details.get("能力目录五级", ""),
           ability_details.get("上架日期", ""),
           ability_details.get("更新日期", ""),
           ability_details.get("能力亮点1", ""),
           ability_details.get("能力亮点2", ""),
           ability_details.get("能力亮点3", ""),
           ability_details.get("能力亮点4", ""),
           ability_details.get("能力亮点5", ""),
           ability_details.get("售前电话", ""),
           ability_details.get("技术支持电话", ""),
           ability_details.get("客服电话", "")]
    
    for i, row_in_ws in enumerate(ability_details_list_ws.iter_rows(min_row=2, max_row=ability_details_list_ws.max_row, values_only=True)):
        # 如果该行为空行，说明到了有实际内容的末尾，在此行写入
        if all(cell is None or str(cell).strip() == "" for cell in row_in_ws):
            for j, _ in enumerate(row_in_ws):
                ability_details_list_ws.cell(row=i+2, column=j+1, value=row[j])
            try:
                wb.save(target_file_name)
            except Exception as e:
                print(f"保存Excel文件时发生错误：{e}")
                print(f"检查文件 {target_file_name} 是否已在其它软件中打开，关闭后重试")
            return
        # 如果该行不为空，且能力名称匹配，则更新该行
        if row_in_ws[0] == ability_details.get("能力名称", ""):
            for j, _ in enumerate(row_in_ws):
                ability_details_list_ws.cell(row=i+2, column=j+1, value=row[j])
            try:
                wb.save(target_file_name)
            except Exception as e:
                print(f"保存Excel文件时发生错误：{e}")
                print(f"检查文件 {target_file_name} 是否已在其它软件中打开，关闭后重试")
            return  # 找到即写入并返回

def browse_ability_list(driver, target_file_name, wb, abilities_count_cell, ability_details_list_ws):
    """
    遍历页面的能力清单项，检查Excel文件中该能力状态是否为"查询成功",是就跳过
    否则点击每一项进入能力详情页面，提取信息
    """
    try:
        # 先获取能力清单页中，搜索到的能力总数
        total_abilities = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((
                By.XPATH, "//div[contains(@class, 'hx-product-left-select-result')]/span"
                ))
        ).text # text返回元素的文本内容
        total_abilities = int(total_abilities.strip())
        print(f"搜索到的能力总数：{total_abilities} 个")

        # 将搜索总数写到目录页
        abilities_count_cell.value = total_abilities
        wb.save(target_file_name)
    except Exception as e:
        print(f"获取能力总数时发生错误：{str(e)}")
        return

    # 获取能力详情列表页能力名称和状态两列
    ability_names_in_sheet = [cell.value for cell in ability_details_list_ws['A'][1:]] # 跳过表头
    ability_statuses_in_sheet = [cell.value for cell in ability_details_list_ws['B'][1:]] # 跳过表头
    # 将能力名称和状态与页面能力项一一对应，生成字典
    ability_status_dict = dict(zip(ability_names_in_sheet, ability_statuses_in_sheet))
    page_num = 1 # 当前页码
    while True:
        print(f"\n浏览能力清单第 {page_num} 页...")
        # 1. 获取当前页所有能力
        try:
            current_page_abilities = WebDriverWait(driver, 15).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "goods-content-title"))
                )
        except TimeoutException:
            print(f"第 {page_num} 页未加载出能力列表，停止浏览")
            break

        # 2. 遍历当前页所有能力项，进入能力详情页面提取信息
        for ability in current_page_abilities:
            # 检查该能力名称在Excel文件中是否存在且状态为"查询成功"，是则跳过
            if ability.text.strip() in ability_status_dict and ability_status_dict[ability.text.strip()] == "查询成功":
                print(f"能力已在Excel中记录，跳过：{ability.text.strip()}")
                continue
            ability_details = {}
            ability_details["能力名称"] = ability.text.strip()
            time.sleep(random.uniform(1, 2))  # 随机等待1-2秒，模拟人类操作
            ability.click()  # 进入能力详情页面
            print(f"\n进入能力详情页面：{ability.text}，提取信息中...")

            # 等待页面加载
            # time.sleep(random.uniform(1, 2))  # 随机等待1-2秒，模拟人类操作
            driver.switch_to.window(driver.window_handles[2])  # selenium切换到新打开的详情页窗口
            # 提取能力项信息
            ability_details = extract_ability_details(driver, ability_details)
            if ability_details and ability_details["状态"] == "查询成功":
                print(f"查询成功：\n{ability_details}")
                # 将能力详情写入Excel
                write_single_ability_detail_to_excel(target_file_name, wb, ability_details_list_ws, ability_details)
            else:
                ability_details_part = {"能力名称": ability.text.strip(), "状态": "查询失败"}
                write_single_ability_detail_to_excel(target_file_name, wb, ability_details_list_ws, ability_details_part)
                print(f"查询失败！")
            # 关闭能力详情页，Selenium切换回能力清单页
            time.sleep(random.uniform(0.5, 1))
            driver.close()
            driver.switch_to.window(driver.window_handles[1])

        # 3. 当前页遍历完，检查是否有下一页可点击
        try:
            next_page_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((
                    By.XPATH, "//li[contains(@class, 'ivu-page-next') and not(contains(@class, 'ivu-page-disabled'))]"
                ))
            )
            time.sleep(random.uniform(1, 2))
            next_page_btn.click() # 点击下一页
            page_num += 1
        except (TimeoutException, ElementNotInteractableException):
            print(f"已浏览所有页面，共{page_num} 页")
            # 关闭能力清单页面，回到首页
            break   
    return


def get_company_info(driver, company_name, target_file_name, wb, abilities_count_cell, ability_details_list_ws):
    """
    单个公司信息检索与提取
    :param company_name: 公司名称
    :param driver: 浏览器驱动实例
    :return: 提取的信息字典
    """
    try:
        # 1. 首页搜索公司名称，进入在新窗口打开的公司能力清单页面
        target_company = search_target_company(driver, company_name)
        if not target_company:
            return {"公司名称": company_name, "状态": "查询失败"}

        time.sleep(random.uniform(0.5, 1))
        # 2.点击找到的公司名称,浏览器自动在新窗口打开能力清单页面
        target_company.click()
        # 将selenium切换到新打开的窗口
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(random.uniform(1, 2))
        print(f"进入公司能力清单页面：")
        # 2.遍历能力清单页面，再进入能力详情页面，提取信息后返回首页
        try:
            browse_ability_list(driver, target_file_name, wb, abilities_count_cell, ability_details_list_ws)

            # 如果查询后abilities_count_cell为0，说明该公司无能力
            if abilities_count_cell.value == 0:
                print(f"公司“{company_name}”查询到的能力数为0，可能尚未上传平台，查询失败！")

            col_b = list(ability_details_list_ws["B"][1:]) # 跳过表头
            valid_values = [cell.value for cell in col_b if cell.value is not None and str(cell.value).strip() != ""]
            if valid_values and all(value == "查询成功" for value in valid_values):
                return {"公司名称": company_name, "状态": "查询成功"}
            elif valid_values and any(value == "查询成功" for value in valid_values):
                return {"公司名称": company_name, "状态": "查询部分成功"}
            else:
                return {"公司名称": company_name, "状态": "查询失败"}

        except Exception as e:
            print(f"能力清单加载异常，停止浏览：{str(e)}")
            return {"公司名称": company_name, "状态": "查询失败"}
        finally:
            # 关闭能力清单页面，回到首页
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

    except (TimeoutException, Exception) as e:
        print(f"获取公司“{company_name}”能力详情时发生错误！")
        print(f"错误信息：{str(e)}")
        return {"公司名称": company_name, "状态": "查询失败"}

def main():
    # 配置参数
    target_url = "https://atom.189.cn/pc/#/index"  # 替换为目标网站的URL
    target_file_path = "01_company_info_by_selenium\\原子能力平台_0919.xlsx"  # 目标文件，至少包含“公司名称”sheet
    driver_path = "chromedriver-win64\\chromedriver.exe"  # 浏览器驱动路径（若已添加到环境变量可省略）
    user_data_dir = "E:\\01_project\\00_task_tool\\chromedriver-win64\\chrome_profile"  # Chrome用户数据目录路径,保存cookie,登录状态等

    # 1. 初始化Chrome浏览器驱动
    options = webdriver.ChromeOptions()
    # 可选：无头模式（不显示浏览器窗口，加快速度）
    # options.add_argument("--headless=new")
    # 保存登录信息
    options.add_argument(f"--user-data-dir={user_data_dir}")
    # 禁用自动化控制特征（部分网站会检测，可选）
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    # 初始化
    driver = webdriver.Chrome(service=ChromeService(driver_path), options=options)
    driver.maximize_window()  # 最大化窗口，避免元素被遮挡

    # 2. 打开网站
    open_and_manual_login(driver, target_url)

    # 检查页面加载状况
    try:
        WebDriverWait(driver, 3).until(
            EC.visibility_of_element_located((By.CLASS_NAME, "username-box"))
        )
    except TimeoutException:
        while True:
            input_str = input("页面加载异常或需要登录，请处理后输入ok继续...\n")
            if input_str.strip().lower() == "ok":
                break

    # 3.读取Excel文件中的目录sheet
    wb = openpyxl.load_workbook(target_file_path)
    dir_ws = wb["目录"]
    template_ws = wb["所需标签"]
    
    # 读取目录页表头，获得公司名称列、能力总数和状态列的索引
    dir_header = [cell.value for cell in dir_ws[2]]  # 第二行是表头
    try:
        company_name_col_idx = dir_header.index("公司名称") 
        abilities_count_col_idx = dir_header.index("能力总数") 
        status_col_idx = dir_header.index("状态")
        timestamp_col_idx = dir_header.index("查询时间")
    except ValueError:
        print("目录sheet缺少“公司名称”或“能力总数”或“状态”或“查询时间”列，请检查！")
        return  # 如果缺少列，终止程序

    # 遍历目录sheet公司名称，从第三行开始
    for _, row in enumerate(dir_ws.iter_rows(min_row=3, values_only=False)):
        company_name_cell = row[company_name_col_idx]
        company_name = company_name_cell.value
        status_cell = row[status_col_idx]
        abilities_count_cell = row[abilities_count_col_idx]
        timestamp_cell= row[timestamp_col_idx]
        if not company_name or str(company_name).strip() == "":
            print("检测到空白公司名称，跳过...")
            continue
        # 如果公司状态为“查询成功”，则跳过
        if status_cell.value == "查询成功":
            print(f"公司“{company_name}”已查询成功，跳过...")
            continue
        # 尝试创建能力详情sheet
        if company_name in wb.sheetnames:
            single_company_ws = wb[company_name]
        else:
            single_company_ws = wb.copy_worksheet(template_ws)
            single_company_ws.title = company_name
            # copy_template(template_ws, single_company_ws) # 字体、边框、保护、合并单元格等都可以复制
        
        # 公司名称添加超链接
        # 链接到本文档中内容，不能直接给hyperlink赋值，这种方式只能链接到文件或网页
        # 需要用到Hyperlink类,并指定ref和location, ref是单元格位置，location是目标位置
        # location格式：sheetname!A1
        # 如果sheetname包含空格、括号或特殊字符，需要用单引号括起来 'sheet name'!A1
        # 因此统一用单引号括起来即可
        company_name_cell.hyperlink = Hyperlink(ref=company_name_cell.coordinate, location=f"'{company_name}'!A1")
        company_name_cell.font = Font(name=company_name_cell.font.name,
                                      size=company_name_cell.font.size,
                                      color=colors.BLUE,
                                      underline="single"
                                      )
        try:
            wb.save(target_file_path)
        except Exception as e:
            print(f"保存Excel文件时发生错误：{e}")
            print(f"检查文件 {target_file_path} 是否已在其它软件中打开，关闭后重试")

        # 获取能力详情并填充到文件中
        company_info = get_company_info(driver, company_name, target_file_path,wb, abilities_count_cell, single_company_ws)

        status_cell.value = company_info["状态"] # 更新公司状态
        timestamp_cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S') # 更新查询时间
        print(f"公司“{company_name}”信息获取完成，状态：{company_info['状态']}")
        try:
            wb.save(target_file_path)
        except Exception as e:
            print(f"保存Excel文件时发生错误：{e}")
            print(f"检查文件 {target_file_path} 是否已在其它软件中打开，关闭后重试")
    # 将目录页设置为活动页面
    wb.active = dir_ws
    try:
        wb.save(target_file_path)
    except Exception as e:
        print(f"保存Excel文件时发生错误：{e}")
        print(f"检查文件 {target_file_path} 是否已在其它软件中打开，关闭后重试")

    print("所有公司已遍历完成！")
    # 关闭浏览器
    driver.quit()

if __name__ == "__main__":
    main()