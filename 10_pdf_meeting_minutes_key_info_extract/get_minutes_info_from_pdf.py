import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import unicodedata
import re
import os
import sys
from datetime import datetime
# from pypdf import PdfReader
from PyPDF2 import PdfReader
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

class PDFExtractorAPP:
    def __init__(self, root):
        self.root = root
        self.root.title("会议纪要PDF信息提取工具")
        self.root.geometry("600x400")

        self.pdf_folder = tk.StringVar()
        self.status_msg = tk.StringVar()
        # 严格和模板字段出现顺序完全一致
        self.fields = [
            'PDF文件名',
            '制发单位', '签发', '编号', '页数',
            '会议名称', '会议时间', '会议地点', '参会人员', '会议主持', '记录人', '会议议题',
            '项目名称', '客户名称', '签约主体', '商机编码', '项目背景', '商机需求', '方案内容',
            '业务模式', '项目金额、成本、利润率', '合同期限', '付款方式', '要求工期',
            '自研自有能力（云中台解决方案）', '方案技术路线可行性（云中台方案+交付/云网小微）',
            '项目运维（云中台方案+交付/云网）', '交付工期（云中台交付+采购/云网）',
            '后项采购方式（云中台交付+采购）', '采购范围与供应链适配性（采购）',
            '结算方式（云中台交付）', '业务场景/模式（政企）', '收入构成及列收计划（政企/财务）',
            '风险点识别等（云中台、政企、法务、财务、采购）',
            '主送', '主办部门', '拟稿人', '核稿人', '发出时间'
        ]
        self._create_ui()

        # 处理文件的路径，考虑打包成exe和直接运行py文件两种情况
        if getattr(sys, 'frozen', False):
            self.basedir  = os.path.dirname(sys.executable)
        else:
            self.basedir = os.path.dirname(os.path.abspath(__file__))


        for file in os.listdir(self.basedir):
            full_path = os.path.join(self.basedir, file)
            if  os.path.isdir(full_path) and "会议纪要PDF" in file:
                self.pdf_folder.set(full_path)
                self._log(f"自动识别到文件夹：{file}")
                break
    
    def _create_ui(self):
        file_frame = tk.LabelFrame(self.root, text="选择文件夹", padx=10, pady=10)
        file_frame.pack(fill='x', padx=10, pady=5)
        tk.Label(file_frame, text='会议纪要：').grid(row=0, column=0, sticky='w')
        tk.Entry(file_frame, textvariable=self.pdf_folder, width=50).grid(row=0, column=1, padx=5)
        tk.Button(file_frame, text='浏览...', command=self._select_dir).grid(row=0, column=2)

        action_frame= tk.Frame(self.root, padx=10, pady=10)
        action_frame.pack(fill='x', padx=10, pady=5)
        tk.Button(
            action_frame,
            text='开始提取',
            command=self._run_extraction,
            bg='#4CAF50',
            fg='white',
            font=('微软雅黑', 10, 'bold')
            ).pack(side='left', padx=5)
        tk.Label(action_frame, textvariable=self.status_msg).pack(side='left', padx=30)

        log_frame = tk.LabelFrame(self.root, text='运行日志', padx=10, pady=10)
        log_frame.pack(fill='both', expand=True, padx=10, pady=5)
        self.log_text = scrolledtext.ScrolledText(log_frame, height=20)
        self.log_text.pack(fill='both', expand=True)


    def _select_dir(self):
        dir_path = filedialog.askdirectory()
        if dir_path:
            self.pdf_folder.set(dir_path)
            self._log(f"已选择文件夹：{dir_path}")
    
    def _log(self, message, level: str = 'info'):
        level_upper = level.upper()
        log_line = f'\n{level_upper} - {message}\n'
        self.log_text.insert(tk.END, log_line)
        self.log_text.see(tk.END)

        self.root.update() # 立即刷新界面显示日志

    def _extract_pdf_text(self, pdf_path):
        def clean_text(text):
            # ------------- 步骤1：保护【正常换行】，替换为临时标记 -------------
            # 保护1：线下/线上 开头的分行（会议地点专用）
            text = re.sub(r'\n(线下|线上)', r'###NEWLINE###\1', text)
            # 保护2：部门/分公司/云中台 开头的分行（参会人员、风险点识别专用）
            text = re.sub(r'\n[一二三四五六七八九十\d、]*\s*(.*?部[)）]*[:：]|.*?分公司|.*?中心[:：]|.*?事业群[:：]|云中台[:：]|政企群[:：]|财务[:：]|采购[:：]|法务[:：])', r'###NEWLINE###\1', text)

            # ------------- 步骤2：仅删除【异常截断换行】（一句话中间的错误换行） -------------
            # 规则：无结束标点 + 换行 + 普通文字 → 去除
            text = re.sub(r'([^。！？；])\n+([^\s])', r'\1\2', text)

            # ------------- 步骤3：恢复所有【正常换行】 -------------
            text = text.replace('###NEWLINE###', '\n')

            # 清理多余空行（保留段落结构）
            text = re.sub(r'\n{3,}', '\n\n', text).strip()
            return text

        try:
            reader = PdfReader(pdf_path)
            full_text = ""
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    full_text += t + "\n"
            return clean_text(full_text)
        except Exception as e:
            print(f"读取PDF失败：{pdf_path}，错误：{e}")
            return ""

    # ===================== 核心提取函数（全需求适配） =====================
    def _extract_field_content(self, text, pdf_name):
        # 格式化时间函数
        def format_meeting_time(time_str):
            time_str = time_str.replace('：', ':').replace(' ', '')
            time = datetime.strptime(time_str, '%Y年%m月%d日%H:%M')
            return time.strftime('%Y年%m月%d日%H:%M')

        def format_issue_time(time_str):
            time_str = time_str.replace('：', ':').replace(' ', '')
            time = datetime.strptime(time_str, '%Y年%m月%d日')
            return time.strftime('%Y年%m月%d日')

        res = {k: "" for k in self.fields}
        res["PDF文件名"] = pdf_name

        # ==============================================
        # 需求1：严格适配模板
        # ==============================================
        base_map = {
            "制发单位": r"制发单位[:：]\s*(.*?)(?=签发|$)",
            "签发": r"签发[:：]\s*(.*?)(?=编号|$)",  # 带冒号匹配
            "编号": r"编号\s*(.*?)(?=密级|$)",
            "页数": r"页数\s*(.*?)(?=会议名称|$)",
            "会议名称": r"会议名称\s*(.*?)(?=会议时间|$)",
            "会议时间": r"会议时间\s*(.*?)(?=会议地点|$)",
            "会议地点": r"会议地点\s*(.*?)(?=参会人员|$)",
            "参会人员": r"参会人员\s*(.*?)(?=会议主持|$)",
            "会议主持": r"会议主持\s*(.*?)(?=记录人|$)",
            "记录人": r"记录人\s*(.*?)(?=会议议题|$)",
            "会议议题": r"会议议题\s*(.*?)(?=会议结论|$)",
            "项目名称": r"项目名称[:：]\s*(.*?)(?=客户名称|$)",
            "客户名称": r"客户名称[:：]\s*(.*?)(?=签约主体|$)",
            "签约主体": r"签约主体[:：]\s*(.*?)(?=商机编码|$)",
            "商机编码": r"商机编码[:：]\s*(.*?)(?=项目背景|$)",
            "项目背景": r"项目背景[:：]\s*(.*?)(?=商机需求|$)",
            "商机需求": r"商机需求[:：]\s*(.*?)(?=方案内容|$)",
            "方案内容": r"方案内容[:：]\s*(.*?)(?=业务模式|$)",
            "业务模式": r"业务模式[:：]\s*(.*?)(?=项目金额|$)",
            "项目金额、成本、利润率": r"(项目金额.*?)(?=合同期限|$)",
            "合同期限": r"合同期限[:：]\s*(.*?)(?=付款方式|$)",
            "付款方式": r"付款方式[:：]\s*(.*?)(?=要求工期|$)",
            "要求工期": r"要求工期[:：]\s*(.*?)(?=。|\n|$)",
            "自研自有能力（云中台解决方案）": r"自研自有能力（.*?[:：]\s*(.*?)(?=方案技术路线可行性|$)",
            "方案技术路线可行性（云中台方案+交付/云网小微）": r"方案技术路线可行性（.*?[:：]\s*(.*?)(?=项目运维|$)",
            "项目运维（云中台方案+交付/云网）": r"项目运维（.*?[:：]\s*(.*?)(?=交付工期|$)",
            "交付工期（云中台交付+采购/云网）": r"交付工期（.*?[:：]\s*(.*?)(?=后项采购方式|$)",
            "后项采购方式（云中台交付+采购）": r"后项采购方式（.*?[:：]\s*(.*?)(?=采购范围|$)",
            "采购范围与供应链适配性（采购）": r"采购范围与供应链适配性（.*?[:：]\s*(.*?)(?=结算方式|$)",
            "结算方式（云中台交付）": r"结算方式（.*?[:：]\s*(.*?)(?=业务场景|$)",
            "业务场景/模式（政企）": r"业务场景/模式（.*?[:：]\s*(.*?)(?=收入构成|$)",
            "收入构成及列收计划（政企/财务）": r"收入构成及列收计划（.*?[:：]\s*(.*?)(?=风险点识别|$)",
            "风险点识别等（云中台、政企、法务、财务、采购）": r"风险点识别等（.*?[:：]\s*(.*?)(?=经会议讨论|$)"
        }
        for key, pat in base_map.items():
            match = re.search(pat, text, re.DOTALL) 
            if match:
                val = match.group(1).strip().rstrip("。.") # 去除末尾的句号等符号
                # 清理参会人员开头的 八、、1.、等垃圾字符
                if key == "参会人员":
                    val = re.sub(r'^[一二三四五六七八九十\d、\s]+', '', val)
                if key == "会议时间":
                    try:
                        val = format_meeting_time(val)
                    except:
                        pass
                res[key] = val

        # ==============================================
        # 底部字段匹配（主送/主办/拟稿/核稿/发出时间）
        # ==============================================
        bottom_map = {
            "主送": r"主送\s*(.*?)(?=抄送|$)",
            "主办部门": r"主办部门\s*(.*?)(?=拟稿人|$)",
            "拟稿人": r"拟稿人\s*(.*?)(?=会签部门|$)",
            "核稿人": r"(?:核稿人|部门核稿)\s*(.*?)(?=会签部分|发出时间|$)",
            "发出时间": r"发出时间\s*(.*)"
        }
        for key, pat in bottom_map.items():
            match = re.search(pat, text, re.DOTALL)
            if match:
                val = match.group(1).strip().rstrip('。.')
                # 清理参会人员和主送开头的 八、、1.、等垃圾字符
                if key == "主送":
                    val = re.sub(r'^[一二三四五六七八九十\d、\s]+', '', val)
                    val = re.sub(r'[\s]+', '', val) # 去除空格、换行
                if key == "发出时间":
                    try:
                        val = format_issue_time(val)
                    except:
                        pass
                res[key] = val

        return res

    # ===================== Excel 美化（专业政企风格，不变） =====================
    def _beautify_excel(self, ws):
        header_fill = PatternFill('solid', fgColor='0070C0') # PattenFill是openpyxl.styles中的类，用于设置单元格的填充样式。'solid'表示实心填充，fgColor='0070C0'表示填充颜色为蓝色
        header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        zebra = 'EBF1F8'
        # 全部框线
        border = Border(left=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

        # 表头样式
        for col in range(1, ws.max_column + 1):
            c = ws.cell(1, col)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border
        ws.row_dimensions[1].height = 30

        # 数据行样式
        for row in range(2, ws.max_row + 1):
            fill = PatternFill('solid', fgColor=zebra if (row-2)%2==0 else 'FFFFFF') # zebra斑马纹，偶数行淡蓝色，奇数行白色
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row, col)
                c.fill, c.border, c.alignment = fill, border,  left
            ws.row_dimensions[row].height = 14.4

        # 列宽自适应
        for col in ws.columns: # ws.columns是一个属性，返回一个生成器，生成工作表中每一列的单元格元组
            letter = col[0].column_letter # col[0]的数据类型是Cell, column_letter是他的属性
            ws.column_dimensions[letter].width = min(50, max(12, max(
                sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(v.value or '')) # unicodedata.east_asian_width(c)是unicodedata模块中的一个函数，用于获取字符c的东亚宽度属性，返回值可能是'F'（全角）、'W'（宽字符）、'A'（半宽字符）、'N'（中性字符）等。这里的意思是，如果是全角或宽字符，则占2个单位，否则占1个单位。
                for v in col if v.value
            )*1.2))
        ws.freeze_panes = 'A2'

    # ===================== 主流程（批量处理PDF） =====================
    def _run_extraction(self):
        folder = self.pdf_folder.get()
        if not folder or not os.path.isdir(folder): # isdir()函数用于判断给定的路径是否是一个存在的目录
            self._log("请选择有效的文件夹路径！", level='error')
            messagebox.showerror("错误", "请选择有效的文件夹路径！")
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 处理结果文件的路径，考虑打包成exe和直接运行py文件两种情况
        output_file = os.path.join(self.basedir, f"会议纪要解析结果_{timestamp}.xlsx")

        all_data = []
        for file in os.listdir(folder):
            if file.lower().endswith('.pdf'):
                path = os.path.join(folder, file)
                self._log(f"正在处理：{file}")
                text = self._extract_pdf_text(path)
                if not text:
                    self._log(f"跳过文件：{file}，无法提取文本内容", level='warning')
                    continue
                data = self._extract_field_content(text, file)
                all_data.append(data)
                self._log(f"处理完成：{file}")
        if not all_data:
            self._log("未找到有效的PDF文件!", level='error')
            messagebox.showerror("错误", "未找到有效的PDF文件!")
            return
        
        # 生成Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(self.fields)

        for d in all_data:
            ws.append([d[f] for f in self.fields])
        self._log(f"全部处理完成！共解析 {len(all_data)} 个PDF文件")
        
        # 美化表格
        self._beautify_excel(ws)
        try:
            wb.save(output_file)
            self._log(f"结果文件已保存：{output_file}")
        except Exception as e:
            self._log(f"保存Excel失败：{output_file}\n{e}", level='error')
            messagebox.showerror("错误", f"保存Excel失败：{output_file}\n{e}")
               
        


if __name__ == '__main__':
    root = tk.Tk()
    app = PDFExtractorAPP(root)
    root.mainloop()